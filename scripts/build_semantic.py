"""
build_semantic.py
=================

Build the semantic mini-wiki page from `handovers/semantic/semantic.json`.

Source-of-truth model (schema_version 2):
    `semantic.json` is the live SoT consumed by both the editor
    (`semantic_editor.html` served by `serve_admin.py`) and the build step.
    `semantic.xlsx` is a human-friendly seed / import / export artifact —
    snapshotted to `<root>/semantic/backup/` before each build or export so
    it stays recoverable side-by-side with the rendered HTML.

Modes:

    python scripts/build_semantic.py --seed
        Create `handovers/semantic/semantic.json` (and a companion seed
        `semantic.xlsx`) with example seed data. Refuses to overwrite an
        existing json.

    python scripts/build_semantic.py --import-xlsx
        Read `semantic.xlsx`, validate, and write `semantic.json`. Use after a
        manual xlsx edit or for a fresh bootstrap when only the xlsx exists.

    python scripts/build_semantic.py --export-xlsx
        Read `semantic.json` and write `semantic.xlsx` (after snapshotting the
        existing xlsx to backup/). Use after editor sessions to refresh the
        human-readable artifact.

    python scripts/build_semantic.py        (default = build)
        Read the JSON, validate (V3 tiered: errors block; warnings render with
        badges), snapshot the current `semantic.xlsx` (if any) to
        `<root>/semantic/backup/` as a timestamped backup, and emit
        `<root>/semantic/08_Semantic_Model.html` as a single-file,
        zero-dependency HTML page. All writes are confined to the current root
        (handovers/ or handovers/source/).

Design references:
    - ADR 0002: XLSX is source of truth for this page. (Superseded by 0005.)
    - ADR 0003: Hand-positioned domains + auto-ring tables.
    - ADR 0004: Curated layer derived from bucket_table_xref.access=W.
    - ADR 0005: semantic.json is SoT; xlsx is seed/export/backup.

See handovers/CONTEXT.md for the canonical vocabulary.
"""
from __future__ import annotations

import argparse
import json
import math
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    print("✗ openpyxl is required. Install with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

# Force UTF-8 stdout so non-ASCII + box-drawing chars don't crash on Windows cp1252.
try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except Exception:
    pass

ROOT = Path(__file__).resolve().parents[1]   # handovers/ (or handovers/source/)
# Source-of-truth JSON + rendered HTML live at <ROOT>/semantic/. The companion
# xlsx is the human-friendly seed/import/export artifact. Before each build the
# current xlsx is snapshotted to <ROOT>/semantic/backup/ as a timestamped
# `.bak.xlsx`. All writes are confined to <ROOT>; this script never reaches
# into a sibling tree (e.g. handovers/source/ vs handovers/) — those are
# independent replicas with their own copy of this script.
JSON_PATH  = ROOT / "semantic" / "semantic.json"
XLSX       = ROOT / "semantic" / "semantic.xlsx"
OUT_HTML   = ROOT / "semantic" / "08_Semantic_Model.html"
BACKUP_DIR = ROOT / "semantic" / "backup"
INV_JSON   = ROOT / "inventory.json"

SCHEMA_VERSION = 2

TABS = ["domains", "tables", "columns", "relationships", "bucket_table_xref", "meta"]

DOMAIN_COLS = ["domain_id", "domain_name", "description", "x", "y", "radius", "color", "icon"]
TABLE_COLS = ["table_fqn", "domain_id", "short_name", "description",
              "grain", "partition_col", "source_system", "share", "notes"]
COLUMN_COLS = ["table_fqn", "column", "type", "description",
               "domain_values", "nullable", "pii", "notes"]
# `from_column` / `to_column` are optional (added in schema_version 2) — they
# carry structured endpoints for column-level FK edges authored via the editor;
# `via` is kept as the authoritative human-readable string and may be derived.
REL_COLS = ["from", "to", "kind", "cardinality", "via", "notes",
            "from_column", "to_column"]
XREF_COLS = ["bucket_id", "table_fqn", "access", "notes"]
META_COLS = ["key", "value"]

REL_KINDS = {"domain_link", "fk", "derived_from"}
ACCESS_KINDS = {"R", "W", "RW"}


# ─── Validation result container ───────────────────────────────────────────

@dataclass
class BuildReport:
    errors: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)

    def err(self, msg: str) -> None:
        self.errors.append(msg)

    def warn(self, msg: str) -> None:
        self.warnings.append(msg)

    @property
    def ok(self) -> bool:
        return not self.errors

    def print_summary(self) -> None:
        print()
        print("─" * 60)
        if self.errors:
            print(f"✗ {len(self.errors)} error(s), {len(self.warnings)} warning(s)")
            for e in self.errors:
                print(f"  ✗ {e}")
        else:
            print(f"✓ 0 errors, {len(self.warnings)} warning(s) — built OK")
        for w in self.warnings:
            print(f"  ⚠ {w}")
        print("─" * 60)


# ─── Seed mode ─────────────────────────────────────────────────────────────

def _seed_data() -> dict[str, list[dict[str, Any]]]:
    """Return a minimal generic seed payload (Acme Retail Analytics demo).

    Two domains, three tables, a handful of cols / rels / xref rows — just
    enough to exercise every render path. Replace wholesale with your team's
    real semantic model via the editor (or by editing semantic.json directly).
    """

    def _d(*vals):
        return dict(zip(DOMAIN_COLS, vals))

    def _t(*vals):
        return dict(zip(TABLE_COLS, vals))

    def _c(*vals):
        return dict(zip(COLUMN_COLS, vals))

    def _r(*vals):
        padded = list(vals) + [""] * (len(REL_COLS) - len(vals))
        return dict(zip(REL_COLS, padded))

    def _x(*vals):
        return dict(zip(XREF_COLS, vals))

    domains = [
        _d("CUSTOMER", "Customer", "Customer master, demographics, churn risk.",        30, 40, 80, "#0EA5E9", "🧑"),
        _d("ORDER",    "Order",    "Order headers, line items, returns, lifecycle.",    70, 40, 80, "#4F46E5", "🛒"),
    ]

    tables = [
        _t("raw_catalog.crm.customers",
           "CUSTOMER", "customers", "Customer master — one row per customer_id (CDC latest state).",
           "1 row per customer_id", "", "CRM", False, ""),
        _t("raw_catalog.commerce.orders",
           "ORDER", "orders", "Order headers from the e-commerce platform.",
           "1 row per order_id", "order_date", "Commerce", False, ""),
        _t("curated_catalog.analytics.customer_360",
           "CUSTOMER", "customer_360",
           "Curated 360° customer profile (demographics + order history + LTV).",
           "1 row per customer_id", "", "DERIVED", True, "Shared downstream — keep schema stable."),
    ]

    columns = [
        _c("raw_catalog.crm.customers",          "customer_id",  "STRING", "Customer ID — primary key.",       "", False, False, ""),
        _c("raw_catalog.crm.customers",          "email",        "STRING", "Customer email (PII).",            "", True,  True,  "PII — handle per data-handling policy."),
        _c("raw_catalog.commerce.orders",        "order_id",     "STRING", "Order ID — primary key.",          "", False, False, ""),
        _c("raw_catalog.commerce.orders",        "customer_id",  "STRING", "FK → customers.customer_id.",      "", False, False, ""),
        _c("curated_catalog.analytics.customer_360", "customer_id", "STRING", "Primary key (joined from raw).", "", False, False, ""),
        _c("curated_catalog.analytics.customer_360", "lifetime_value", "DECIMAL(18,2)", "LTV in account currency.", "", False, False, ""),
    ]

    relationships = [
        _r("CUSTOMER", "ORDER", "domain_link", "1:n", "customers.customer_id = orders.customer_id", ""),
        _r("raw_catalog.crm.customers",
           "raw_catalog.commerce.orders",
           "fk", "1:n", "customers.customer_id = orders.customer_id", "",
           "customer_id", "customer_id"),
        _r("raw_catalog.crm.customers",
           "curated_catalog.analytics.customer_360",
           "derived_from", "1:1", "Curated 360° projection of customer state.", "",
           "customer_id", "customer_id"),
    ]

    xref = [
        _x("PRJ-2026-01", "raw_catalog.crm.customers",                  "R", "Source for demographics."),
        _x("PRJ-2026-01", "raw_catalog.commerce.orders",                "R", "Source for order history."),
        _x("PRJ-2026-01", "curated_catalog.analytics.customer_360",     "W", "Pipeline writes the 360° table."),
    ]

    meta = [
        {"key": "schema_version", "value": SCHEMA_VERSION},
        {"key": "last_updated", "value": ""},
        {"key": "author", "value": ""},
        {"key": "notes", "value": "Seeded by build_semantic.py --seed. Replace/extend with your team's domains."},
    ]

    return {
        "domains": domains,
        "tables": tables,
        "columns": columns,
        "relationships": relationships,
        "bucket_table_xref": xref,
        "meta": meta,
    }


def _seed() -> None:
    """Bootstrap `semantic.json` (SoT) and a companion `semantic.xlsx`.

    Refuses to overwrite an existing json. The xlsx is only written if it
    doesn't already exist — protecting any hand-curated xlsx from being
    clobbered. Run `--export-xlsx` later to refresh the xlsx from json.
    """
    if JSON_PATH.exists():
        print(f"✗ {JSON_PATH} already exists. Refusing to overwrite.", file=sys.stderr)
        print("  Delete or rename it first if you really want to re-seed.", file=sys.stderr)
        sys.exit(2)

    JSON_PATH.parent.mkdir(parents=True, exist_ok=True)
    data = _seed_data()
    _write_json(data)
    print(f"✓ Seeded {JSON_PATH.relative_to(ROOT.parent)}")

    if not XLSX.exists():
        _write_xlsx(data)
        print(f"✓ Seeded companion {XLSX.relative_to(ROOT.parent)}")
    else:
        print(f"  (skipped {XLSX.name} — already exists; run --export-xlsx to refresh)")

    print(f"  Open the editor with: python handovers/serve_admin.py  (then /semantic_editor.html)")
    print(f"  Or build the wiki directly: python handovers/scripts/build_semantic.py")


# ─── Read mode ─────────────────────────────────────────────────────────────

def _read_xlsx() -> dict[str, list[dict[str, Any]]]:
    """Load every tab as a list of dicts keyed by header row. Missing tabs → []."""
    wb = load_workbook(XLSX, data_only=True, read_only=True)
    out: dict[str, list[dict[str, Any]]] = {t: [] for t in TABS}
    for tab in TABS:
        if tab not in wb.sheetnames:
            continue
        ws = wb[tab]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(c).strip() if c is not None else "" for c in rows[0]]
        for r in rows[1:]:
            if r is None or all(c is None or (isinstance(c, str) and not c.strip()) for c in r):
                continue
            out[tab].append({h: r[i] if i < len(r) else None for i, h in enumerate(headers)})
    # Schema v2 back-fill: derive from_column/to_column from `via` text when
    # the xlsx pre-dates v2 and the via string matches the canonical
    # "<table>.<col_a> = <table>.<col_b>" pattern.
    _backfill_rel_columns(out.get("relationships", []))
    return out


# ─── Via-string helpers (kind=fk endpoint round-trip) ──────────────────────

import re as _re

_VIA_PATTERN = _re.compile(
    r"^\s*([\w\.\$]+?)\.(\w+)\s*=\s*([\w\.\$]+?)\.(\w+)\s*$"
)


def _parse_via(via: str) -> tuple[str | None, str | None]:
    """Parse a `via` string of form 'tbl_a.col_a = tbl_b.col_b' and return
    (from_column, to_column). Returns (None, None) if it doesn't match."""
    if not via or not isinstance(via, str):
        return None, None
    m = _VIA_PATTERN.match(via)
    if not m:
        return None, None
    return m.group(2), m.group(4)


def _backfill_rel_columns(rels: list[dict[str, Any]]) -> None:
    """Mutate each relationship row in-place: if from_column/to_column are
    missing/blank and `via` matches the canonical pattern, fill them in."""
    for r in rels:
        if not r.get("from_column") or not r.get("to_column"):
            fc, tc = _parse_via(r.get("via") or "")
            if fc and not r.get("from_column"):
                r["from_column"] = fc
            if tc and not r.get("to_column"):
                r["to_column"] = tc


def _derive_via(rel: dict[str, Any]) -> str:
    """Given a relationship row with structured from/to/from_column/to_column,
    return a canonical `via` string. Falls back to whatever `via` already
    contains if structured fields aren't both present, or if `from`/`to`
    aren't fully-qualified table names (e.g. domain_link rows use domain ids)."""
    fc, tc = rel.get("from_column"), rel.get("to_column")
    f, t = rel.get("from"), rel.get("to")
    if fc and tc and f and t and "." in str(f) and "." in str(t):
        # Use the bare table name (last segment) for readability.
        f_tbl = str(f).rsplit(".", 1)[-1]
        t_tbl = str(t).rsplit(".", 1)[-1]
        return f"{f_tbl}.{fc} = {t_tbl}.{tc}"
    return rel.get("via") or ""


# ─── JSON I/O (primary SoT after schema_version 2) ─────────────────────────

def _read_json() -> dict[str, list[dict[str, Any]]]:
    """Load `semantic.json` into the same shape `_read_xlsx` returns:
    {tab_name: [row_dict, ...], ...}. Missing tabs → []."""
    raw = json.loads(JSON_PATH.read_text(encoding="utf-8"))
    out: dict[str, list[dict[str, Any]]] = {t: [] for t in TABS}
    for tab in TABS:
        rows = raw.get(tab) or []
        if not isinstance(rows, list):
            rows = []
        out[tab] = [dict(r) for r in rows if isinstance(r, dict)]
    _backfill_rel_columns(out.get("relationships", []))
    return out


def _write_json(data: dict[str, list[dict[str, Any]]]) -> None:
    """Persist the authoring dict to `semantic.json`. Derives `via` from
    structured fields where possible to keep the human-readable string in
    sync. Wraps with a top-level schema_version + tabs payload."""
    rels = data.get("relationships") or []
    for r in rels:
        derived = _derive_via(r)
        if derived:
            r["via"] = derived
    payload = {"schema_version": SCHEMA_VERSION}
    for tab in TABS:
        payload[tab] = data.get(tab, [])
    JSON_PATH.parent.mkdir(parents=True, exist_ok=True)
    JSON_PATH.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )


def _write_xlsx(data: dict[str, list[dict[str, Any]]]) -> None:
    """Persist the authoring dict to `semantic.xlsx`. Header row + one row per
    record. Used by --seed (when no xlsx exists), --export-xlsx, and
    `serve_admin.py`'s POST /export-semantic-xlsx endpoint."""
    column_order = {
        "domains":           DOMAIN_COLS,
        "tables":            TABLE_COLS,
        "columns":           COLUMN_COLS,
        "relationships":     REL_COLS,
        "bucket_table_xref": XREF_COLS,
        "meta":              META_COLS,
    }
    wb = Workbook()
    wb.remove(wb.active)
    for tab in TABS:
        ws = wb.create_sheet(tab)
        cols = column_order[tab]
        ws.append(cols)
        for row in (data.get(tab) or []):
            ws.append([row.get(c) for c in cols])
    XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(XLSX)


def _backup_xlsx() -> str | None:
    """Snapshot the current `semantic.xlsx` (if any) to
    `<ROOT>/semantic/backup/semantic.<ts>.bak.xlsx`. Returns the backup path
    relative to ROOT.parent for logging, or None if there was nothing to back
    up. Task-1 entry point: replaces the prior HTML-backup logic."""
    if not XLSX.exists():
        return None
    import shutil
    from datetime import datetime
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d-%H%M%S")
    backup_path = BACKUP_DIR / f"{XLSX.stem}.{ts}.bak.xlsx"
    shutil.copy2(XLSX, backup_path)
    try:
        return f"{backup_path.relative_to(ROOT.parent)} ({backup_path.stat().st_size:,} bytes)"
    except ValueError:
        return f"{backup_path} ({backup_path.stat().st_size:,} bytes)"


def _read_inventory_buckets() -> set[str]:
    """Return the set of valid bucket_ids from inventory.json (for xref validation)."""
    if not INV_JSON.exists():
        return set()
    try:
        data = json.loads(INV_JSON.read_text(encoding="utf-8"))
        return {b.get("bucket_id") for b in data.get("buckets", []) if b.get("bucket_id")}
    except Exception:
        return set()


# ─── Validate (V3 tiered) ──────────────────────────────────────────────────

def _validate(data: dict[str, list[dict[str, Any]]], rpt: BuildReport) -> None:
    domains  = data["domains"]
    tables   = data["tables"]
    columns  = data["columns"]
    rels     = data["relationships"]
    xref     = data["bucket_table_xref"]

    # ERRORS — these would render a misleading graph

    if not domains:
        rpt.err("tab 'domains' is empty — at least one domain is required.")
    if not tables:
        rpt.err("tab 'tables' is empty — at least one table is required.")

    seen_dom: set[str] = set()
    for i, d in enumerate(domains, start=2):
        did = d.get("domain_id")
        if not did:
            rpt.err(f"domains row {i}: missing domain_id."); continue
        if did in seen_dom:
            rpt.err(f"domains row {i}: duplicate domain_id '{did}'.")
        seen_dom.add(did)
        if not d.get("domain_name"):
            rpt.err(f"domains row {i} ({did}): missing domain_name.")
        if not d.get("description"):
            rpt.err(f"domains row {i} ({did}): missing description.")
        for ax in ("x", "y"):
            v = d.get(ax)
            if v is None:
                rpt.err(f"domains row {i} ({did}): missing {ax}.")
            else:
                try:
                    fv = float(v)
                    if not (0 <= fv <= 100):
                        rpt.err(f"domains row {i} ({did}): {ax}={v} not in 0–100.")
                except (TypeError, ValueError):
                    rpt.err(f"domains row {i} ({did}): {ax}={v!r} is not numeric.")

    seen_tbl: set[str] = set()
    for i, t in enumerate(tables, start=2):
        fqn = t.get("table_fqn")
        if not fqn:
            rpt.err(f"tables row {i}: missing table_fqn."); continue
        if fqn in seen_tbl:
            rpt.err(f"tables row {i}: duplicate table_fqn '{fqn}'.")
        seen_tbl.add(fqn)
        if fqn.count(".") < 1:
            rpt.err(f"tables row {i} ({fqn}): table_fqn should be catalog.schema.table.")
        did = t.get("domain_id")
        if not did:
            rpt.err(f"tables row {i} ({fqn}): missing domain_id.")
        elif did not in seen_dom:
            rpt.err(f"tables row {i} ({fqn}): domain_id '{did}' not in domains tab.")
        for req in ("short_name", "description"):
            if not t.get(req):
                rpt.err(f"tables row {i} ({fqn}): missing {req}.")

    for i, r in enumerate(rels, start=2):
        if r.get("kind") not in REL_KINDS:
            rpt.err(f"relationships row {i}: kind={r.get('kind')!r} not in {sorted(REL_KINDS)}.")
        for end in ("from", "to"):
            v = r.get(end)
            if not v:
                rpt.err(f"relationships row {i}: missing {end}.")

    for i, x in enumerate(xref, start=2):
        if not x.get("bucket_id"):
            rpt.err(f"bucket_table_xref row {i}: missing bucket_id.")
        if not x.get("table_fqn"):
            rpt.err(f"bucket_table_xref row {i}: missing table_fqn.")
        if x.get("access") not in ACCESS_KINDS:
            rpt.err(f"bucket_table_xref row {i}: access={x.get('access')!r} not in {sorted(ACCESS_KINDS)}.")

    # WARNINGS — incomplete but not misleading

    documented_tables: set[str] = {c.get("table_fqn") for c in columns if c.get("table_fqn")}
    for fqn in seen_tbl:
        if fqn not in documented_tables:
            rpt.warn(f"table '{fqn}' has no rows in the 'columns' tab (data dictionary missing).")

    xref_tables: set[str] = {x.get("table_fqn") for x in xref if x.get("table_fqn")}
    for fqn in xref_tables - seen_tbl:
        rpt.warn(f"bucket_table_xref references unknown table '{fqn}' (not in 'tables' tab).")

    inv_buckets = _read_inventory_buckets()
    if inv_buckets:
        for x in xref:
            bid = x.get("bucket_id")
            if bid and bid not in inv_buckets:
                rpt.warn(f"bucket_table_xref bucket_id '{bid}' not found in inventory.json buckets.")

    # Undocumented curated tables (have W-row but share=False/missing)
    written_by: dict[str, list[str]] = {}
    read_by:    dict[str, list[str]] = {}
    for x in xref:
        fqn, bid, acc = x.get("table_fqn"), x.get("bucket_id"), x.get("access")
        if not fqn or not bid:
            continue
        if acc in ("W", "RW"):
            written_by.setdefault(fqn, []).append(bid)
        if acc in ("R", "RW"):
            read_by.setdefault(fqn, []).append(bid)

    share_flag = {t.get("table_fqn"): bool(t.get("share")) for t in tables}
    for fqn, writers in written_by.items():
        readers = [b for b in read_by.get(fqn, []) if b not in writers]
        if readers and not share_flag.get(fqn):
            rpt.warn(f"curated table '{fqn}' is read by {readers} but share=FALSE — set share=TRUE to surface it on the map.")

    # Orphan domains (no tables)
    used_doms = {t.get("domain_id") for t in tables}
    for did in seen_dom - used_doms:
        rpt.warn(f"domain '{did}' has no tables.")


# ─── Render (single-file HTML) ─────────────────────────────────────────────

def _layer_for(fqn: str, written_by: dict[str, list[str]]) -> str:
    return "curated" if fqn in written_by else "raw"


def _render(data: dict[str, list[dict[str, Any]]]) -> str:
    domains = data["domains"]
    tables  = data["tables"]
    columns = data["columns"]
    rels    = data["relationships"]
    xref    = data["bucket_table_xref"]
    meta    = {m.get("key"): m.get("value") for m in data["meta"]}

    # Aggregate xref
    written_by: dict[str, list[str]] = {}
    read_by:    dict[str, list[str]] = {}
    for x in xref:
        fqn, bid, acc = x.get("table_fqn"), x.get("bucket_id"), x.get("access")
        if not fqn or not bid:
            continue
        if acc in ("W", "RW"):
            written_by.setdefault(fqn, []).append(bid)
        if acc in ("R", "RW"):
            read_by.setdefault(fqn, []).append(bid)

    # Build the runtime JSON payload the front-end consumes.
    domain_records = []
    for d in domains:
        domain_records.append({
            "id": d.get("domain_id"),
            "name": d.get("domain_name"),
            "description": d.get("description") or "",
            "x": float(d.get("x") or 50),
            "y": float(d.get("y") or 50),
            "radius": int(d.get("radius") or 80),
            "color": d.get("color") or "#6366F1",
            "icon": d.get("icon") or "",
        })

    columns_by_table: dict[str, list[dict[str, Any]]] = {}
    for c in columns:
        fqn = c.get("table_fqn")
        if not fqn:
            continue
        columns_by_table.setdefault(fqn, []).append({
            "column": c.get("column"),
            "type": c.get("type"),
            "description": c.get("description") or "",
            "domain_values": c.get("domain_values") or "",
            "nullable": bool(c.get("nullable")),
            "pii": bool(c.get("pii")),
            "notes": c.get("notes") or "",
        })

    table_records = []
    # Determine which curated tables earn a bubble (share=True OR shared across buckets)
    share_flag = {t.get("table_fqn"): bool(t.get("share")) for t in tables}
    for t in tables:
        fqn = t.get("table_fqn")
        layer = _layer_for(fqn, written_by)
        # Curated tables: only render if share=True OR (multiple distinct buckets touch it)
        # Raw tables: always render.
        all_buckets = set(written_by.get(fqn, []) + read_by.get(fqn, []))
        render_it = (
            layer == "raw"
            or share_flag.get(fqn)
            or len(all_buckets) >= 2
        )
        table_records.append({
            "fqn": fqn,
            "domain_id": t.get("domain_id"),
            "short_name": t.get("short_name"),
            "description": t.get("description") or "",
            "grain": t.get("grain") or "",
            "partition_col": t.get("partition_col") or "",
            "source_system": t.get("source_system") or "",
            "layer": layer,
            "share": share_flag.get(fqn, False),
            "notes": t.get("notes") or "",
            "render": render_it,
            "written_by": sorted(set(written_by.get(fqn, []))),
            "read_by":    sorted(set(read_by.get(fqn, []))),
            "columns":    columns_by_table.get(fqn, []),
            "documented": fqn in columns_by_table,
        })

    rel_records = [{
        "from": r.get("from"),
        "to": r.get("to"),
        "kind": r.get("kind"),
        "cardinality": r.get("cardinality") or "",
        "via": r.get("via") or "",
        "notes": r.get("notes") or "",
        "from_column": r.get("from_column") or "",
        "to_column": r.get("to_column") or "",
    } for r in rels]

    payload = {
        "schema_version": SCHEMA_VERSION,
        "meta": meta,
        "domains": domain_records,
        "tables": table_records,
        "relationships": rel_records,
    }

    payload_json = json.dumps(payload, ensure_ascii=False, indent=2)

    # Embed payload via a script tag with type="application/json" so the HTML
    # is single-file and the front-end can `JSON.parse` it on load.
    return _HTML_TEMPLATE.replace("__PAYLOAD__", payload_json)


# ─── HTML template (single file, zero external deps) ──────────────────────

_HTML_TEMPLATE = r"""<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<title>Semantic Model — VN AI Wiki</title>
<meta name="viewport" content="width=device-width,initial-scale=1">
<style>
  /* ─── Birchline palette (mirrors handovers/assets/wiki.css) ──────────── */
  :root {
    --ivory:  #FAF9F5;
    --paper:  #FFFFFF;
    --slate:  #141413;
    --clay:   #D97757;
    --clay-d: #B85C3E;
    --oat:    #E3DACC;
    --olive:  #788C5D;
    --crimson:#A8453A;
    --info:   #5C7CA3;
    --g100:   #F0EEE6;
    --g200:   #E6E3DA;
    --g300:   #D1CFC5;
    --g500:   #6F6E68;
    --g700:   #3D3D3A;
    --shadow-1: 0 1px 2px rgba(20,20,19,.06);
    --shadow-2: 0 4px 10px rgba(20,20,19,.08);
    --shadow-3: 0 12px 28px rgba(20,20,19,.12);
    --ease:   cubic-bezier(.16, 1, .3, 1);
    --spring: cubic-bezier(.34, 1.56, .64, 1);
    --sans:   system-ui, -apple-system, "Segoe UI", Roboto, Helvetica, Arial, sans-serif;
    --mono:   ui-monospace, "SF Mono", Menlo, Monaco, Consolas, monospace;
    /* semantic-page specific */
    --raw:     var(--info);     /* raw / source tables */
    --curated: var(--clay);     /* team-owned / curated tables */
    --edge:    var(--g300);
    --edge-hi: var(--clay-d);
    --pii:     var(--crimson);
  }
  * { box-sizing: border-box; }
  html, body { margin: 0; height: 100%; background: var(--ivory); color: var(--slate);
    font-family: var(--sans); line-height: 1.5; -webkit-font-smoothing: antialiased; }
  a { color: var(--clay-d); text-decoration: none; }
  a:hover { text-decoration: underline; }

  header { padding: 12px 24px; display: flex; align-items: center; gap: 18px; flex-wrap: wrap;
    background: var(--paper); border-bottom: 1.5px solid var(--g300); box-shadow: var(--shadow-1); }
  header h1 { font-size: 17px; margin: 0; font-weight: 600; color: var(--slate); }
  header .crumbs { font-family: var(--mono); font-size: 11px; letter-spacing: .08em;
    text-transform: uppercase; color: var(--g500); }
  header .crumbs a { color: var(--clay-d); }
  .legend { display: flex; gap: 14px; font-size: 12px; color: var(--g500); align-items: center; }
  .legend .swatch { display: inline-block; width: 10px; height: 10px; border-radius: 50%;
    vertical-align: middle; margin-right: 6px; }
  #reset-layout { display: none; margin-left: 6px; padding: 3px 9px;
    background: var(--paper); color: var(--g500); border: 1px solid var(--g300);
    border-radius: 6px; font-family: var(--sans); font-size: 11px; cursor: pointer;
    transition: background .15s var(--ease), color .15s var(--ease); }
  #reset-layout:hover { background: var(--clay-l); color: var(--clay-d); border-color: var(--clay); }

  /* ─── Search omnibox ────────────────────────────────────────────────── */
  .search-wrap { position: relative; flex: 1 1 280px; max-width: 460px;
    margin: 0 6px; }
  #search { width: 100%; padding: 7px 32px 7px 32px; font-size: 13px; font-family: var(--sans);
    border: 1.5px solid var(--g300); border-radius: 8px; background: var(--ivory);
    color: var(--slate); outline: none; transition: border-color .2s, box-shadow .2s; }
  #search::placeholder { color: var(--g500); }
  #search:focus { border-color: var(--clay); box-shadow: 0 0 0 3px rgba(217,119,87,.18);
    background: var(--paper); }
  .search-wrap::before { content: "⌕"; position: absolute; left: 10px; top: 50%;
    transform: translateY(-50%); color: var(--g500); font-size: 14px; pointer-events: none; }
  .search-kbd { position: absolute; right: 8px; top: 50%; transform: translateY(-50%);
    font-family: var(--mono); font-size: 10px; color: var(--g500);
    background: var(--g100); border: 1px solid var(--g200); border-radius: 4px;
    padding: 1px 5px; pointer-events: none; }
  #search:focus ~ .search-kbd { display: none; }

  #search-results { position: absolute; left: 0; right: 0; top: calc(100% + 6px);
    background: var(--paper); border: 1.5px solid var(--g300); border-radius: 10px;
    box-shadow: var(--shadow-3); max-height: 70vh; overflow-y: auto;
    z-index: 50; padding: 6px 0; display: none; }
  #search-results.open { display: block; }
  #search-results .sr-group { padding: 4px 0; }
  #search-results .sr-group + .sr-group { border-top: 1px solid var(--g200); }
  #search-results .sr-head { padding: 6px 14px 4px; font-family: var(--mono);
    font-size: 10px; letter-spacing: .08em; text-transform: uppercase; color: var(--g500); }
  #search-results .sr-row { padding: 6px 14px; cursor: pointer; font-size: 13px;
    display: flex; align-items: baseline; gap: 8px; line-height: 1.35; }
  #search-results .sr-row:hover, #search-results .sr-row.active { background: var(--g100); }
  #search-results .sr-name { color: var(--slate); font-family: var(--mono);
    word-break: break-all; }
  #search-results .sr-name b { color: var(--clay-d); font-weight: 700; background: rgba(217,119,87,.12); }
  #search-results .sr-meta { color: var(--g500); font-size: 11px; font-family: var(--mono);
    margin-left: auto; flex-shrink: 0; max-width: 55%; text-align: right;
    word-break: break-all; }
  #search-results .sr-empty { padding: 14px; color: var(--g500); font-style: italic;
    font-size: 12px; text-align: center; }
  #search-results .sr-more { padding: 4px 14px 6px; font-size: 11px; color: var(--g500);
    font-style: italic; }

  /* Flash highlight on a column row after a column search-result click */
  @keyframes col-flash {
    0%   { background: rgba(217,119,87,.35); }
    100% { background: transparent; }
  }
  table.cols tr.col-flash > td { animation: col-flash 1.6s var(--ease); }

  .legend .swatch { display: inline-block; width: 10px; height: 10px; border-radius: 50%;
    vertical-align: middle; margin-right: 6px; }

  main { display: grid; grid-template-columns: minmax(0, 1fr) 400px; height: calc(100vh - 54px); }
  #canvas { position: relative; overflow: hidden; background:
    radial-gradient(1200px 800px at 30% 20%, var(--paper) 0%, var(--ivory) 70%); }
  svg { width: 100%; height: 100%; display: block; }

  /* ─── View toggle (Domain map / Table model) ──────────────────────── */
  .view-toggle { display: inline-flex; border: 1.5px solid var(--g300);
    border-radius: 999px; overflow: hidden; background: var(--paper);
    margin-left: 12px; font-family: var(--mono); font-size: 12px; }
  .view-toggle button { background: transparent; color: var(--g700); border: 0;
    padding: 5px 12px; cursor: pointer; font: inherit; letter-spacing: .02em; }
  .view-toggle button.active { background: var(--clay); color: var(--paper); }
  .view-toggle button:hover:not(.active) { background: var(--g100); }

  /* ─── Zoom toolbar (Domain map mode) ──────────────────────────────── */
  .zoom-toolbar { position: absolute; top: 12px; right: 12px; z-index: 5;
    display: inline-flex; border: 1.5px solid var(--g300); border-radius: 8px;
    overflow: hidden; background: var(--paper); box-shadow: var(--shadow-1);
    font-family: var(--mono); font-size: 12px; }
  .zoom-toolbar button { background: var(--paper); border: 0; color: var(--g700);
    padding: 5px 10px; cursor: pointer; font: inherit; min-width: 30px;
    border-right: 1px solid var(--g200); }
  .zoom-toolbar button:last-child { border-right: 0; }
  .zoom-toolbar button:hover { background: var(--g100); color: var(--clay-d); }
  .zoom-toolbar .zoom-readout { padding: 5px 10px; color: var(--g500);
    border-right: 1px solid var(--g200); min-width: 44px; text-align: center; }
  #canvas.is-panning { cursor: grabbing !important; }
  #canvas.is-panning .domain-bubble { cursor: grabbing; }

  /* ─── Table model mode ────────────────────────────────────────────── */
  /* Hidden by default; revealed when the user picks Table model. We reuse the
     existing #side renderer's table-card markup, just laid out as a grid and
     given an HTML-level zoom (CSS transform: scale on the inner container). */
  #table-model { position: absolute; inset: 0; overflow: auto; display: none;
    background: var(--ivory); padding: 24px; }
  body.view-table #table-model { display: block; }
  body.view-table #svg { display: none; }
  body.view-table aside#side { display: none; }
  body.view-table main { grid-template-columns: 1fr; }
  #table-model .tm-inner { transform-origin: 0 0;
    transition: transform .2s var(--ease); }
  #table-model .tm-domain { margin: 0 0 30px; }
  #table-model .tm-domain-head { display: flex; align-items: center; gap: 10px;
    margin: 0 0 12px; padding: 6px 12px; border-radius: 999px;
    background: var(--paper); border: 1.5px solid var(--g300); width: max-content;
    font-family: var(--mono); font-weight: 600; font-size: 14px; }
  #table-model .tm-domain-head .icn { font-size: 18px; }
  #table-model .tm-domain-head .swatch { width: 10px; height: 10px; border-radius: 50%; }
  #table-model .tm-grid { display: grid; gap: 14px;
    grid-template-columns: repeat(auto-fill, minmax(280px, 1fr)); }
  #table-model .table-card { margin: 0; max-height: 480px; overflow: auto;
    transition: outline .2s var(--ease), box-shadow .25s var(--ease); }
  #table-model .table-card.flash { outline: 2px solid var(--clay);
    outline-offset: 3px; box-shadow: 0 0 0 6px rgba(217,119,87,.18); }
  #table-model .tm-empty { color: var(--g500); font-style: italic;
    padding: 40px; text-align: center; }

  /* ─── Search highlight on satellite bubbles ───────────────────────── */
  /* When the user picks a Table or Column result from the omnibox, the matching
     satellite gets `.search-hit` (sticky highlight) and all OTHER currently-
     visible satellites get `.search-dim` so the eye lands on the right one even
     when the domain has 20+ tables ringed around it. Cleared on next select. */
  .table-bubble.search-hit circle.tbl { stroke: var(--clay); stroke-width: 3;
    fill: var(--clay-l);
    filter: drop-shadow(0 0 14px var(--clay)); }
  .table-bubble.search-hit text { fill: var(--clay-d); font-weight: 700; }
  .table-bubble.search-dim { opacity: .25; }
  @keyframes sat-pulse {
    0%, 100% { transform: scale(1); }
    35%      { transform: scale(1.45); }
    70%      { transform: scale(1.18); }
  }
  .table-bubble.search-hit.visible { animation: sat-pulse 1.6s var(--spring) 1; }


  /* ─── Domain bubble (big, hand-positioned) ────────────────────────── */
  /* ─── Domain bubble (big, hand-positioned) ────────────────────────── */
  /* The CSS `transform` composes a per-domain nudge (CSS vars --ndx/--ndy,
   * set by the drag system) with the existing :hover / .active scale and
   * the breathe animation. SVG `transform=` attribute is reserved for
   * satellites + edges (which don't share these CSS animations). */
  .domain-bubble { cursor: grab;
    transition: transform .35s var(--ease), filter .35s var(--ease);
    transform-origin: center; transform-box: fill-box;
    transform: translate(var(--sdx,0px), var(--sdy,0px)) translate(var(--ndx,0px), var(--ndy,0px)); }
  .domain-bubble.dragging { cursor: grabbing; transition: none; }
  .domain-bubble.dragging,
  .domain-bubble.dragging:hover { transform: translate(var(--sdx,0px), var(--sdy,0px)) translate(var(--ndx,0px), var(--ndy,0px)); }
  .domain-bubble circle.dom { fill: var(--paper); stroke-width: 3;
    filter: drop-shadow(0 6px 14px rgba(20,20,19,.10)); }
  .domain-bubble text.label { fill: var(--slate); font-size: 15px; font-weight: 600;
    text-anchor: middle; pointer-events: none; font-family: var(--sans); }
  .domain-bubble text.icon  { font-size: 26px; text-anchor: middle; pointer-events: none; }
  .domain-bubble text.count { fill: var(--g500); font-size: 11px; text-anchor: middle;
    pointer-events: none; font-family: var(--mono); letter-spacing: .04em; }
  .domain-bubble:hover { transform: translate(var(--sdx,0px), var(--sdy,0px)) translate(var(--ndx,0px), var(--ndy,0px)) scale(1.06); }
  .domain-bubble:hover circle.dom { filter: drop-shadow(0 0 18px var(--clay)); }
  .domain-bubble.active { transform: translate(var(--sdx,0px), var(--sdy,0px)) translate(var(--ndx,0px), var(--ndy,0px)) scale(1.12); }
  .domain-bubble.active circle.dom { filter: drop-shadow(0 0 22px var(--clay)); }

  @keyframes breathe {
    0%,100% { transform: translate(var(--sdx,0px), var(--sdy,0px)) translate(var(--ndx,0px), var(--ndy,0px)) scale(1); }
    50%     { transform: translate(var(--sdx,0px), var(--sdy,0px)) translate(var(--ndx,0px), var(--ndy,0px)) scale(1.02); }
  }
  .domain-bubble { animation: breathe 5s ease-in-out infinite; }
  .domain-bubble:nth-of-type(2n) { animation-delay: -1.5s; }
  .domain-bubble:nth-of-type(3n) { animation-delay: -3s; }
  .domain-bubble.active, .domain-bubble:hover, .domain-bubble.dragging { animation: none; }

  /* ─── Satellite table bubble (small, ring-positioned) ──────────────── */
  /* Outer <g> carries the SVG `transform="translate(x,y)"` attribute (positioning).
     Inner <g class="table-bubble"> carries the CSS transform (scale animation).
     Keeping these on separate elements prevents CSS transform from clobbering
     the SVG transform attribute (which would dump bubbles at origin 0,0). */
  .table-bubble { cursor: pointer; opacity: 0;
    transform: scale(0.25);
    transform-origin: center; transform-box: fill-box;
    transition: opacity .35s var(--ease), transform .4s var(--spring); }
  .table-bubble.visible { opacity: 1; transform: scale(1); }
  .table-bubble circle.tbl { fill: var(--paper); stroke: var(--raw); stroke-width: 2;
    filter: drop-shadow(0 2px 4px rgba(20,20,19,.08)); }
  .table-bubble.curated circle.tbl { stroke: var(--curated); stroke-dasharray: 5 3; }
  .table-bubble text { fill: var(--slate); font-size: 10px; text-anchor: middle;
    pointer-events: none; font-family: var(--mono); }
  .table-bubble:hover circle.tbl { fill: var(--g100);
    filter: drop-shadow(0 0 10px var(--clay)); }

  /* ─── Dense-domain dots ─────────────────────────────────────────────── */
  /* When a domain owns more than 12 satellites, the labels overlap and the
     ring becomes a hairball. We collapse those satellites into tiny dots
     in BOTH overview and Solo focus, mirroring how the Semantic Editor
     renders crowded domains. The reader reveals one table at a time by
     hovering its dot — the dot grows and its label appears. Search hits
     also enlarge the corresponding dot. There is deliberately no
     "expand all" affordance: click-to-solo positions the focused puck in
     the centre but does NOT re-inflate the satellites, because doing so
     reproduces exactly the overlapping-labels failure we are trying to
     avoid. */
  .satellites-wrap.dense .table-bubble circle.tbl { r: 3.5; stroke-width: 1.5; }
  .satellites-wrap.dense .table-bubble text       { display: none; }
  .satellites-wrap.dense .table-bubble:hover circle.tbl { r: 22; stroke-width: 2; }
  .satellites-wrap.dense .table-bubble:hover text       { display: block; }
  .satellites-wrap.dense .table-bubble.search-hit circle.tbl { r: 6; }
  .satellites-wrap.dense .table-bubble.search-hit text       { display: block; }

  /* ─── Edges ─────────────────────────────────────────────────────────── */
  .edge { stroke: var(--edge); stroke-width: 1.5; fill: none; opacity: .6;
    transition: opacity .3s, stroke .3s, stroke-width .3s; }
  .edge.kind-derived_from { stroke-dasharray: 5 4; }
  .edge.hi { stroke: var(--edge-hi); opacity: 1; stroke-width: 2.5; }
  .edge.solo-hidden { display: none; }

  /* ─── Solo (focus) mode ─────────────────────────────────────────────── */
  /* A click on a domain enters Solo: focused puck translates to the canvas
     centre via --sdx/--sdy; the other domains slide to a perimeter ring as
     dimmed dashed "ghost" pucks; their satellites are hidden; non-focused
     edges drop out. Esc / empty-click / breadcrumb / re-click exit. */
  .satellites-wrap.solo-hidden { display: none; }
  body.view-solo .domain-bubble.is-ghost {
    opacity: .55; animation: none; pointer-events: auto; cursor: pointer; }
  body.view-solo .domain-bubble.is-ghost:hover { opacity: .85; }
  body.view-solo .domain-bubble.is-ghost circle.dom {
    stroke-dasharray: 4 4; stroke-width: 2; filter: none; }
  body.view-solo .domain-bubble.is-ghost text.count { display: none; }
  body.view-solo .domain-bubble.is-focused { animation: none; }

  #solo-controls { position: absolute; top: 14px; left: 14px;
    display: none; gap: 6px; z-index: 5; pointer-events: auto; }
  body.view-solo #solo-controls { display: inline-flex; }
  body.view-table #solo-controls { display: none; }
  .solo-chip { display: inline-flex; align-items: center; gap: 6px;
    padding: 4px 10px; border-radius: 13px;
    background: var(--paper); border: 1.5px solid var(--g300);
    font-size: 12px; color: var(--g700); cursor: pointer;
    transition: background .15s, border-color .15s, color .15s; }
  .solo-chip:hover { background: var(--clay-l); border-color: var(--clay); color: var(--clay-d); }
  .solo-chip-name { background: transparent; border: 0; padding: 4px 6px;
    color: var(--slate); font-weight: 600; cursor: default; }
  .solo-chip-name:hover { background: transparent; border: 0; color: var(--slate); }
  @keyframes solo-pop {
    0% { opacity: 0; transform: translateY(-4px); }
    100% { opacity: 1; transform: translateY(0); } }
  body.view-solo #solo-controls { animation: solo-pop .25s var(--ease); }

  /* ─── Side panel ────────────────────────────────────────────────────── */
  #side { background: var(--paper); border-left: 1.5px solid var(--g300);
    overflow-y: auto; overflow-x: hidden; padding: 22px 24px; min-width: 0; }
  #side h2 { margin: 0 0 6px; font-size: 17px; font-weight: 600; }
  #side .desc { color: var(--g700); font-size: 13px; margin: 0 0 14px; }
  #side .desc ul { padding-left: 18px; margin: 8px 0; }
  #side .empty { color: var(--g500); font-style: italic; }

  .chip { display: inline-block; padding: 2px 9px; border-radius: 12px; font-size: 11px;
    background: var(--g100); color: var(--g700); margin: 2px 4px 2px 0;
    font-family: var(--mono); border: 1px solid var(--g200); }
  .chip a { color: inherit; text-decoration: none; }
  .chip a:hover { text-decoration: underline; }
  .chip.layer-raw     { background: rgba(92,124,163,.12); color: var(--info); border-color: rgba(92,124,163,.3); }
  .chip.layer-curated { background: rgba(217,119,87,.12); color: var(--clay-d); border-color: rgba(217,119,87,.3); }
  .chip.bucket { background: var(--paper); border-color: var(--clay); color: var(--clay-d); }

  .table-card { background: var(--ivory); border-radius: 10px; padding: 14px 16px;
    margin: 12px 0; border-left: 3px solid var(--raw);
    box-shadow: var(--shadow-1); min-width: 0; }
  .table-card.curated { border-left-color: var(--curated); }
  .table-card.undocumented { opacity: .8; }
  .table-card h3 { margin: 0 0 4px; font-size: 13px; font-weight: 600;
    font-family: var(--mono);
    word-break: break-all; overflow-wrap: anywhere;
    display: flex; align-items: center; gap: 8px; flex-wrap: wrap; }
  .table-card .fqn { color: var(--g500); font-size: 11px; font-family: var(--mono);
    word-break: break-all; overflow-wrap: anywhere; display: block;
    margin-bottom: 8px; }
  .table-card .body { font-size: 13px; color: var(--g700);
    word-break: break-word; overflow-wrap: anywhere; }
  .table-card .meta { color: var(--g500); font-size: 11px; margin-top: 6px;
    word-break: break-word; overflow-wrap: anywhere; }
  .table-card .meta code { background: var(--g100); padding: 1px 5px; border-radius: 3px;
    font-size: 10px; }
  .table-card details { margin-top: 10px; }
  .table-card summary { cursor: pointer; color: var(--clay-d); font-size: 12px;
    font-weight: 500; padding: 4px 0; }
  table.cols { width: 100%; border-collapse: collapse; margin-top: 8px; font-size: 11px;
    table-layout: fixed; }
  table.cols th, table.cols td { padding: 5px 6px; border-bottom: 1px solid var(--g200);
    text-align: left; vertical-align: top;
    word-break: break-word; overflow-wrap: anywhere; }
  table.cols th { color: var(--g500); font-weight: 500; font-family: var(--mono);
    font-size: 10px; text-transform: uppercase; letter-spacing: .05em; }
  table.cols code { background: var(--g100); padding: 1px 4px; border-radius: 3px; }
  .pii { color: var(--pii); }
  .nodoc { color: var(--g500); font-style: italic; margin-top: 10px; font-size: 12px; }

  @media (prefers-reduced-motion: reduce) {
    .domain-bubble, .table-bubble, .edge { animation: none !important; transition: none !important; }
  }
  @media (max-width: 900px) {
    main { grid-template-columns: 1fr; height: auto; }
    #canvas { height: 60vh; }
  }
</style>
</head>
<body>
<header>
  <div class="crumbs"><a href="../index.html">← VN AI Wiki</a> · <a href="../00_SEMANTIC_WALKTHROUGH.html" title="Semantic model walkthrough — what this page is and how to read it">📖 Walkthrough</a></div>
  <h1>🗺️ Semantic Model</h1>
  <div class="search-wrap">
    <input id="search" type="search" autocomplete="off" spellcheck="false"
           placeholder="Search domains, tables, columns…"
           aria-label="Search domains, tables, and columns"
           aria-controls="search-results" aria-expanded="false">
    <span class="search-kbd">/</span>
    <div id="search-results" role="listbox" aria-label="Search results"></div>
  </div>
  <div class="legend">
    <span><span class="swatch" style="background:var(--raw)"></span>Raw (source)</span>
    <span><span class="swatch" style="background:var(--curated)"></span>Curated (team-owned)</span>
    <span id="meta-info"></span>
    <div class="view-toggle" role="tablist" aria-label="View mode">
      <button id="vt-domain" type="button" role="tab" aria-selected="true" class="active" title="Domain map — hand-positioned bubbles with ringed tables">🌐 Domain map</button>
      <button id="vt-table"  type="button" role="tab" aria-selected="false" title="Table model — every table laid out as a card grid">🧱 Table model</button>
    </div>
    <button id="reset-layout" type="button" title="Clear any drag nudges saved in this browser. The baked layout is unchanged.">↺ Reset layout</button>
  </div>
</header>
<main>
  <div id="canvas">
    <svg id="svg" viewBox="0 0 1000 700" preserveAspectRatio="xMidYMid meet"></svg>
    <div id="solo-controls" role="group" aria-label="Solo focus controls">
      <button id="solo-back" type="button" class="solo-chip" title="Back to all domains (Esc)">← All domains</button>
      <span id="solo-name" class="solo-chip solo-chip-name"></span>
    </div>
    <div class="zoom-toolbar" id="zoom-toolbar" role="toolbar" aria-label="Zoom">
      <button id="zoom-out"  type="button" title="Zoom out (− or scroll down)" aria-label="Zoom out">−</button>
      <span class="zoom-readout" id="zoom-readout">100%</span>
      <button id="zoom-in"   type="button" title="Zoom in (+ or scroll up)" aria-label="Zoom in">+</button>
      <button id="zoom-fit"  type="button" title="Fit to view (0)" aria-label="Fit to view">⊡</button>
    </div>
    <div id="table-model" aria-hidden="true">
      <div class="tm-inner" id="tm-inner"></div>
    </div>
  </div>
  <aside id="side">
    <div id="welcome">
      <h2>Welcome 👋</h2>
      <p class="desc">Click a domain bubble to expand its tables and view the data dictionary.</p>
      <p class="desc"><b>Tip:</b> drag a domain to nudge it for clarity — saved per-browser, the source layout is untouched.</p>
      <p class="desc"><b>Reading the map:</b></p>
      <ul class="desc">
        <li>Big bubbles are <b>subject-area domains</b> (Policy, Customer, …).</li>
        <li>Small satellite bubbles are <b>physical tables</b>. Solid border = raw source; dashed clay = curated/team-owned.</li>
        <li>Edges show <b>relationships</b> (FKs, lineage). Hover an edge to see the join predicate.</li>
        <li>Each table shows <b>which handovers buckets read/write it</b> — click a bucket chip to jump to its card in the main wiki.</li>
      </ul>
    </div>
    <div id="domain-view" hidden></div>
  </aside>
</main>

<script id="payload" type="application/json">
__PAYLOAD__
</script>
<script>
(() => {
  const data = JSON.parse(document.getElementById('payload').textContent);
  document.getElementById('meta-info').textContent =
    (data.meta && data.meta.last_updated ? 'Updated ' + data.meta.last_updated : '');

  const svg = document.getElementById('svg');
  const W = 1000, H = 700;
  const ns = 'http://www.w3.org/2000/svg';
  const el = (n, a={}) => { const e = document.createElementNS(ns, n);
    for (const k in a) e.setAttribute(k, a[k]); return e; };

  const domById = Object.fromEntries(data.domains.map(d => [d.id, d]));
  const tablesByDom = {};
  for (const t of data.tables) {
    if (!t.render) continue;
    (tablesByDom[t.domain_id] ||= []).push(t);
  }
  const tblByFqn = Object.fromEntries(data.tables.map(t => [t.fqn, t]));
  const domCoord = d => ({ cx: d.x/100 * W, cy: d.y/100 * H });

  // ─── Solo (focus) mode state ─────────────────────────────────────────
  // soloFocusId === null  → overview (current default).
  // soloFocusId === '<id>' → that domain is centred; others are ghost pucks.
  // Per-table solo offsets are read from meta.solo_offsets (a JSON string
  // populated by the Semantic Editor) — we honour them but cannot write back.
  const GHOST_PUCK_INSET = 110;
  let soloFocusId = null;
  let metaSoloOffsets = {};
  try {
    const raw = (data.meta && data.meta.solo_offsets) || '';
    if (raw) metaSoloOffsets = JSON.parse(raw) || {};
  } catch (_) { metaSoloOffsets = {}; }
  function soloOffsetOf(id) {
    // Translation (dx,dy) to add to a domain's natural position when in Solo.
    if (!soloFocusId) return { dx: 0, dy: 0 };
    const d = domById[id]; if (!d) return { dx: 0, dy: 0 };
    const nat = domCoord(d);
    if (id === soloFocusId) {
      // Pull the focused puck to the canvas centre.
      return { dx: W/2 - nat.cx, dy: H/2 - nat.cy };
    }
    // Ghost pucks: project the other domain along the ray from canvas centre
    // through its natural position onto an inset rectangle. Preserves the
    // model's spatial intuition (north stays north, east stays east).
    const focusNat = domCoord(domById[soloFocusId]);
    let vx = nat.cx - focusNat.cx, vy = nat.cy - focusNat.cy;
    if (vx === 0 && vy === 0) { vx = 1; vy = 0; }
    const halfW = W/2 - GHOST_PUCK_INSET;
    const halfH = H/2 - GHOST_PUCK_INSET;
    const sx = halfW / Math.abs(vx || 1e-9);
    const sy = halfH / Math.abs(vy || 1e-9);
    const t = Math.min(sx, sy);
    const targetX = W/2 + vx * t;
    const targetY = H/2 + vy * t;
    return { dx: targetX - nat.cx, dy: targetY - nat.cy };
  }

  // Edges layer first (drawn below bubbles)
  const gEdges = el('g', {id:'edges'});
  svg.appendChild(gEdges);
  for (const r of data.relationships) {
    let a, b, fromDom, toDom;
    if (domById[r.from] && domById[r.to]) {
      a = domCoord(domById[r.from]); b = domCoord(domById[r.to]);
      fromDom = r.from; toDom = r.to;
    } else if (tblByFqn[r.from] && tblByFqn[r.to]) {
      fromDom = tblByFqn[r.from].domain_id;
      toDom   = tblByFqn[r.to].domain_id;
      a = domCoord(domById[fromDom]);
      b = domCoord(domById[toDom]);
      if (!a || !b || a===b) continue;
    } else continue;
    const line = el('line', {x1:a.cx, y1:a.cy, x2:b.cx, y2:b.cy, class:'edge kind-'+r.kind});
    line.dataset.from = r.from; line.dataset.to = r.to;
    // Base coords + endpoint domains — used by the drag system to recompute
    // (x1,y1,x2,y2) when a domain bubble is nudged.
    line.dataset.fromDomain = fromDom; line.dataset.toDomain = toDom;
    line.dataset.bx1 = a.cx; line.dataset.by1 = a.cy;
    line.dataset.bx2 = b.cx; line.dataset.by2 = b.cy;
    gEdges.appendChild(line);
  }

  // Satellites layer — separate so they overlay edges but render under domain bubbles.
  const gSatellites = el('g', {id:'satellites'});
  svg.appendChild(gSatellites);

  // Domain bubbles layer (on top)
  const gDomains = el('g', {id:'domains'});
  svg.appendChild(gDomains);

  for (const d of data.domains) {
    const {cx, cy} = domCoord(d);
    const r = d.radius || 80;

    // Domain bubble
    const g = el('g', {class: 'domain-bubble', 'data-id': d.id});
    g.appendChild(el('circle', {cx, cy, r, class:'dom', stroke: d.color}));
    if (d.icon) {
      const ic = el('text', {x:cx, y:cy-6, class:'icon'}); ic.textContent = d.icon;
      g.appendChild(ic);
    }
    const tName = el('text', {x:cx, y:cy+22, class:'label'}); tName.textContent = d.name;
    g.appendChild(tName);
    const tCount = el('text', {x:cx, y:cy+38, class:'count'});
    tCount.textContent = (tablesByDom[d.id]||[]).length + ' tables';
    g.appendChild(tCount);
    g.addEventListener('click', (ev) => { ev.stopPropagation(); selectDomain(d.id); });
    gDomains.appendChild(g);

    // Satellites — outer <g> for SVG transform (position), inner <g> for CSS transform (scale anim)
    const tables = tablesByDom[d.id] || [];
    const ringR = r + 60;
    const DENSE_THRESHOLD = 12;   // > this → render satellites as dots (always, overview + solo)
    const isDense = tables.length > DENSE_THRESHOLD;
    const wrap = el('g', {class: 'satellites-wrap' + (isDense ? ' dense' : ''),
                          'data-domain': d.id, style:'pointer-events:none'});
    tables.forEach((t, i) => {
      const ang = (-Math.PI/2) + (i * 2 * Math.PI / Math.max(tables.length, 1));
      const tx = cx + ringR * Math.cos(ang);
      const ty = cy + ringR * Math.sin(ang);
      const pos = el('g', {transform: `translate(${tx},${ty})`});  // position only
      pos.dataset.baseTransform = `translate(${tx},${ty})`;
      const tb  = el('g', {class: 'table-bubble' + (t.layer==='curated'?' curated':''),
                           'data-fqn': t.fqn});                     // CSS scale anim
      tb.appendChild(el('circle', {cx:0, cy:0, r:28, class:'tbl'}));
      const fullName = t.short_name || t.fqn.split('.').pop();
      const tip = el('title'); tip.textContent = fullName;
      tb.appendChild(tip);
      appendWrappedLabel(tb, fullName);
      tb.addEventListener('click', (ev) => { ev.stopPropagation(); focusTable(t.fqn); });
      pos.appendChild(tb);
      wrap.appendChild(pos);
    });
    gSatellites.appendChild(wrap);
  }

  // ─── Domain drag (view-only nudges, persisted to localStorage) ─────
  // Lets the reader shuffle overlapping domains apart for clarity. Does NOT
  // touch semantic.json — the canonical layout is baked at build time.
  const NUDGE_KEY = 'vn-aiwiki-domain-nudges-v1';
  let nudges = {};
  try {
    const raw = JSON.parse(localStorage.getItem(NUDGE_KEY) || '{}');
    for (const id in raw) if (domById[id] && raw[id] && typeof raw[id].dx === 'number' && typeof raw[id].dy === 'number') nudges[id] = raw[id];
  } catch (_) { nudges = {}; }
  function nudgeOf(id){ return nudges[id] || {dx:0, dy:0}; }
  function updateResetVisibility(){
    const btn = document.getElementById('reset-layout');
    if (btn) btn.style.display = Object.keys(nudges).length ? 'inline-block' : 'none';
  }
  function persistNudges(){
    try { localStorage.setItem(NUDGE_KEY, JSON.stringify(nudges)); } catch (_) {}
    updateResetVisibility();
  }
  function applyDomainTransform(id){
    const n = nudgeOf(id);
    const s = soloOffsetOf(id);
    // Domain bubbles use CSS vars so the translate composes with the
    // breathe/hover/active scale transforms. Satellites have no such CSS
    // animations and use the plain SVG transform attribute.
    document.querySelectorAll(`.domain-bubble[data-id="${CSS.escape(id)}"]`).forEach(g => {
      g.style.setProperty('--ndx', n.dx + 'px');
      g.style.setProperty('--ndy', n.dy + 'px');
      g.style.setProperty('--sdx', s.dx + 'px');
      g.style.setProperty('--sdy', s.dy + 'px');
    });
    document.querySelectorAll(`.satellites-wrap[data-domain="${CSS.escape(id)}"]`).forEach(g => g.setAttribute('transform', `translate(${n.dx + s.dx},${n.dy + s.dy})`));
    // Per-table solo offsets (from meta.solo_offsets, written by the editor).
    // Applied only to the soloed domain; other domains always reset to base.
    (tablesByDom[id] || []).forEach(t => {
      const tb = document.querySelector(`.satellites-wrap[data-domain="${CSS.escape(id)}"] .table-bubble[data-fqn="${CSS.escape(t.fqn)}"]`);
      if (!tb) return;
      const wrapper = tb.parentNode;
      const base = wrapper && wrapper.dataset.baseTransform;
      if (!base) return;
      const off = (soloFocusId === id) ? metaSoloOffsets[id + '\u0001' + t.fqn] : null;
      wrapper.setAttribute('transform', off ? `${base} translate(${off.x||0},${off.y||0})` : base);
    });
  }
  function applyEdgeOffsetsFor(idSet){
    document.querySelectorAll('.edge').forEach(line => {
      const fd = line.dataset.fromDomain, td = line.dataset.toDomain;
      if (idSet && !(idSet.has(fd) || idSet.has(td))) return;
      const fn = nudgeOf(fd), tn = nudgeOf(td);
      const sfn = soloOffsetOf(fd), stn = soloOffsetOf(td);
      line.setAttribute('x1', (+line.dataset.bx1) + fn.dx + sfn.dx);
      line.setAttribute('y1', (+line.dataset.by1) + fn.dy + sfn.dy);
      line.setAttribute('x2', (+line.dataset.bx2) + tn.dx + stn.dx);
      line.setAttribute('y2', (+line.dataset.by2) + tn.dy + stn.dy);
      // Visibility filter: in Solo, hide edges that don't touch the focused domain.
      const hidden = !!soloFocusId && fd !== soloFocusId && td !== soloFocusId;
      line.classList.toggle('solo-hidden', hidden);
    });
  }
  function applyAllNudges(){
    Object.keys(nudges).forEach(applyDomainTransform);
    applyEdgeOffsetsFor(null);
    updateResetVisibility();
  }
  applyAllNudges();

  let drag = null;
  const DRAG_THRESHOLD = 4;
  function svgPointFromEvent(e){
    const pt = svg.createSVGPoint();
    pt.x = e.clientX; pt.y = e.clientY;
    const ctm = svg.getScreenCTM();
    return ctm ? pt.matrixTransform(ctm.inverse()) : {x: e.clientX, y: e.clientY};
  }
  svg.addEventListener('pointerdown', (e) => {
    const bubble = e.target.closest && e.target.closest('.domain-bubble');
    if (!bubble) return;
    if (e.button !== 0) return;
    const id = bubble.dataset.id;
    const n = nudgeOf(id);
    const p = svgPointFromEvent(e);
    drag = {
      id, bubble,
      startUx: p.x, startUy: p.y,
      baseDx: n.dx, baseDy: n.dy,
      idSet: new Set([id]),
      moved: false,
    };
    try { bubble.setPointerCapture(e.pointerId); } catch (_) {}
    bubble.classList.add('dragging');
  });
  svg.addEventListener('pointermove', (e) => {
    if (!drag) return;
    const p = svgPointFromEvent(e);
    const dx = p.x - drag.startUx;
    const dy = p.y - drag.startUy;
    if (!drag.moved){
      const rect = svg.getBoundingClientRect();
      const sx = rect.width  / svg.viewBox.baseVal.width;
      const sy = rect.height / svg.viewBox.baseVal.height;
      if (Math.hypot(dx*sx, dy*sy) < DRAG_THRESHOLD) return;
      drag.moved = true;
    }
    nudges[drag.id] = {dx: drag.baseDx + dx, dy: drag.baseDy + dy};
    applyDomainTransform(drag.id);
    applyEdgeOffsetsFor(drag.idSet);
  });
  function endDrag(e){
    if (!drag) return;
    const wasMoved = drag.moved;
    drag.bubble.classList.remove('dragging');
    try { drag.bubble.releasePointerCapture(e.pointerId); } catch (_) {}
    drag = null;
    if (wasMoved){
      persistNudges();
      // Swallow the click that follows pointerup so selectDomain doesn't fire.
      const suppress = (ev) => { ev.stopPropagation(); ev.preventDefault(); svg.removeEventListener('click', suppress, true); };
      svg.addEventListener('click', suppress, true);
    }
  }
  svg.addEventListener('pointerup', endDrag);
  svg.addEventListener('pointercancel', endDrag);

  function resetLayout(){
    nudges = {};
    try { localStorage.removeItem(NUDGE_KEY); } catch (_) {}
    // Re-apply per-domain transforms; honours any active Solo state.
    data.domains.forEach(d => applyDomainTransform(d.id));
    applyEdgeOffsetsFor(null);
    updateResetVisibility();
  }
  const resetBtn = document.getElementById('reset-layout');
  if (resetBtn) resetBtn.addEventListener('click', resetLayout);
  updateResetVisibility();

  /* ─── Pan + zoom (Domain map mode) ──────────────────────────────────
   * Wraps every existing top-level <g> inside <svg> in a single viewport
   * <g id="viewport"> and applies a matrix transform. Wheel zooms toward
   * the cursor; drag on empty canvas pans. Click handlers on bubbles still
   * fire because the viewport <g> doesn't intercept pointer events. */
  const canvasEl = document.getElementById('canvas');
  let suppressNextClick = false;
  const viewport = el('g', {id: 'viewport'});
  while (svg.firstChild) viewport.appendChild(svg.firstChild);
  svg.appendChild(viewport);
  const VIEW_VB = { w: 1000, h: 700 };  // matches the SVG viewBox
  const Z_MIN = 0.4, Z_MAX = 5;
  let zoom = 1, panX = 0, panY = 0;
  const readout = document.getElementById('zoom-readout');
  function applyViewport() {
    viewport.setAttribute('transform', `translate(${panX},${panY}) scale(${zoom})`);
    if (readout) readout.textContent = Math.round(zoom * 100) + '%';
  }
  function clientToVB(clientX, clientY) {
    const r = svg.getBoundingClientRect();
    // CSS pixels → viewBox units (preserveAspectRatio=xMidYMid meet)
    const scale = Math.min(r.width / VIEW_VB.w, r.height / VIEW_VB.h);
    const offX = (r.width  - VIEW_VB.w * scale) / 2;
    const offY = (r.height - VIEW_VB.h * scale) / 2;
    return { x: (clientX - r.left - offX) / scale,
             y: (clientY - r.top  - offY) / scale };
  }
  function zoomAt(clientX, clientY, factor) {
    const z2 = Math.max(Z_MIN, Math.min(Z_MAX, zoom * factor));
    if (z2 === zoom) return;
    const p = clientToVB(clientX, clientY);
    // Keep the point under the cursor stable: pan' = p - (p - pan) * z2/z
    panX = p.x - (p.x - panX) * (z2 / zoom);
    panY = p.y - (p.y - panY) * (z2 / zoom);
    zoom = z2;
    applyViewport();
  }
  function fitView(){ zoom = 1; panX = 0; panY = 0; applyViewport(); }
  applyViewport();
  document.getElementById('zoom-in') .addEventListener('click', () => {
    const r = svg.getBoundingClientRect();
    zoomAt(r.left + r.width/2, r.top + r.height/2, 1.25);
  });
  document.getElementById('zoom-out').addEventListener('click', () => {
    const r = svg.getBoundingClientRect();
    zoomAt(r.left + r.width/2, r.top + r.height/2, 1/1.25);
  });
  document.getElementById('zoom-fit').addEventListener('click', fitView);
  svg.addEventListener('wheel', (e) => {
    if (document.body.classList.contains('view-table')) return;
    e.preventDefault();
    zoomAt(e.clientX, e.clientY, e.deltaY < 0 ? 1.12 : 1/1.12);
  }, { passive: false });

  // Pan when dragging on empty canvas (not on a domain bubble or satellite).
  let panning = null;
  svg.addEventListener('pointerdown', (e) => {
    if (e.button !== 0) return;
    if (e.target.closest('.domain-bubble') || e.target.closest('.table-bubble')) return;
    panning = { sx: e.clientX, sy: e.clientY, px: panX, py: panY, moved: false };
    svg.setPointerCapture(e.pointerId);
    canvasEl.classList.add('is-panning');
  });
  svg.addEventListener('pointermove', (e) => {
    if (!panning) return;
    const r = svg.getBoundingClientRect();
    const scale = Math.min(r.width / VIEW_VB.w, r.height / VIEW_VB.h);
    const dx = (e.clientX - panning.sx) / scale;
    const dy = (e.clientY - panning.sy) / scale;
    if (Math.abs(dx) + Math.abs(dy) > 2) panning.moved = true;
    panX = panning.px + dx;
    panY = panning.py + dy;
    applyViewport();
  });
  function endPan(e){
    if (!panning) return;
    try { svg.releasePointerCapture(e.pointerId); } catch(_) {}
    const moved = panning.moved;
    panning = null;
    canvasEl.classList.remove('is-panning');
    // Suppress the click that fires after a real drag (which otherwise dismisses the selection).
    if (moved) { suppressNextClick = true; setTimeout(() => suppressNextClick = false, 0); }
  }
  svg.addEventListener('pointerup',    endPan);
  svg.addEventListener('pointercancel', endPan);

  document.addEventListener('keydown', (e) => {
    if (e.target.matches('input, textarea')) return;
    if (e.key === '+' || e.key === '=') { e.preventDefault();
      const r = svg.getBoundingClientRect(); zoomAt(r.left+r.width/2, r.top+r.height/2, 1.25); }
    else if (e.key === '-' || e.key === '_') { e.preventDefault();
      const r = svg.getBoundingClientRect(); zoomAt(r.left+r.width/2, r.top+r.height/2, 1/1.25); }
    else if (e.key === '0') { e.preventDefault(); fitView(); }
  });

  /* ─── View toggle (Domain map / Table model) ──────────────────────── */
  let tableModelBuilt = false;
  function buildTableModel() {
    const inner = document.getElementById('tm-inner');
    let html = '';
    for (const d of data.domains) {
      const tables = tablesByDom[d.id] || [];
      if (!tables.length) continue;
      html += `<section class="tm-domain" data-domain="${escapeAttr(d.id)}">
        <div class="tm-domain-head">
          <span class="swatch" style="background:${escapeAttr(d.color || '#999')}"></span>
          <span class="icn">${escapeHtml(d.icon || '')}</span>
          <span>${escapeHtml(d.name)}</span>
          <span style="color:var(--g500);font-weight:400">· ${tables.length} tables</span>
        </div>
        <div class="tm-grid">`;
      for (const t of tables) {
        html += `<div class="table-card ${t.layer} ${t.documented?'':'undocumented'}" data-tm-fqn="${escapeAttr(t.fqn)}">
          <h3>${escapeHtml(t.short_name)} <span class="chip layer-${t.layer}">${t.layer}</span></h3>
          <code class="fqn">${escapeHtml(t.fqn)}</code>
          <div class="body">${escapeHtml(t.description || '')}</div>
          ${t.columns && t.columns.length ? `<details><summary>📖 ${t.columns.length} columns</summary>
            <table class="cols">
              <colgroup><col style="width:38%"><col style="width:22%"><col style="width:40%"></colgroup>
              <thead><tr><th>Column</th><th>Type</th><th>Description</th></tr></thead>
              <tbody>${t.columns.map(c => `<tr data-tm-col="${escapeAttr(t.fqn + '::' + c.column)}">
                <td><code>${escapeHtml(c.column)}</code>${c.pii?' <span class="pii">🔒</span>':''}</td>
                <td>${escapeHtml(c.type || '')}</td>
                <td>${escapeHtml(c.description || '')}</td>
              </tr>`).join('')}</tbody>
            </table></details>` : '<div class="nodoc">📝 Not yet documented.</div>'}
        </div>`;
      }
      html += `</div></section>`;
    }
    if (!html) html = '<div class="tm-empty">No tables in this model yet.</div>';
    inner.innerHTML = html;
    tableModelBuilt = true;
  }
  let tmZoom = 1;
  function applyTmZoom(){ document.getElementById('tm-inner').style.transform = `scale(${tmZoom})`;
    if (document.body.classList.contains('view-table')) readout.textContent = Math.round(tmZoom * 100) + '%'; }
  function setView(mode) {
    if (mode === 'table' && !tableModelBuilt) buildTableModel();
    if (mode === 'table') exitSolo();
    document.body.classList.toggle('view-table', mode === 'table');
    document.getElementById('vt-domain').classList.toggle('active', mode === 'domain');
    document.getElementById('vt-table' ).classList.toggle('active', mode === 'table');
    document.getElementById('vt-domain').setAttribute('aria-selected', String(mode === 'domain'));
    document.getElementById('vt-table' ).setAttribute('aria-selected', String(mode === 'table'));
    if (mode === 'table') applyTmZoom(); else applyViewport();
  }
  document.getElementById('vt-domain').addEventListener('click', () => setView('domain'));
  document.getElementById('vt-table' ).addEventListener('click', () => setView('table'));
  // Zoom buttons also drive the Table model when active.
  const _origZoomIn  = document.getElementById('zoom-in');
  const _origZoomOut = document.getElementById('zoom-out');
  const _origZoomFit = document.getElementById('zoom-fit');
  _origZoomIn.addEventListener('click', () => {
    if (!document.body.classList.contains('view-table')) return;
    tmZoom = Math.min(2.5, tmZoom * 1.15); applyTmZoom();
  });
  _origZoomOut.addEventListener('click', () => {
    if (!document.body.classList.contains('view-table')) return;
    tmZoom = Math.max(0.5, tmZoom / 1.15); applyTmZoom();
  });
  _origZoomFit.addEventListener('click', () => {
    if (!document.body.classList.contains('view-table')) return;
    tmZoom = 1; applyTmZoom();
  });

  /* ─── Search highlight on satellite (Domain map) ──────────────────── */
  function clearSearchHighlight() {
    document.querySelectorAll('.table-bubble.search-hit, .table-bubble.search-dim')
      .forEach(b => b.classList.remove('search-hit', 'search-dim'));
    document.querySelectorAll('#tm-inner .table-card.flash')
      .forEach(c => c.classList.remove('flash'));
  }
  function highlightSatelliteByFqn(fqn) {
    const target = document.querySelector(`.table-bubble[data-fqn="${CSS.escape(fqn)}"]`);
    if (!target) return;
    const wrap = target.closest('.satellites-wrap');
    if (!wrap) { target.classList.add('search-hit'); return; }
    // Apply regardless of .visible — selectDomain reveals satellites with a
    // staggered setTimeout (up to ~35ms × n) so we'd otherwise race the reveal.
    wrap.querySelectorAll('.table-bubble').forEach(b => {
      if (b === target) b.classList.add('search-hit');
      else b.classList.add('search-dim');
    });
  }
  function highlightTableModelCard(fqn) {
    const c = document.querySelector(`#tm-inner [data-tm-fqn="${CSS.escape(fqn)}"]`);
    if (!c) return;
    c.classList.add('flash');
    c.scrollIntoView({ behavior: 'smooth', block: 'center' });
    setTimeout(() => c.classList.remove('flash'), 2500);
  }


  function enterSolo(id) {
    if (!domById[id]) return;
    soloFocusId = id;
    document.body.classList.add('view-solo');
    document.querySelectorAll('.domain-bubble').forEach(b => {
      const same = b.dataset.id === id;
      b.classList.toggle('is-focused', same);
      b.classList.toggle('is-ghost', !same);
    });
    const nameEl = document.getElementById('solo-name');
    if (nameEl) nameEl.textContent = (domById[id].icon || '') + ' ' + domById[id].name;
    // Recompute all transforms + edge endpoints/visibility under solo.
    data.domains.forEach(d => applyDomainTransform(d.id));
    applyEdgeOffsetsFor(null);
  }

  function exitSolo() {
    if (!soloFocusId) return;
    soloFocusId = null;
    document.body.classList.remove('view-solo');
    document.querySelectorAll('.domain-bubble').forEach(b => {
      b.classList.remove('is-focused', 'is-ghost');
    });
    // Restore baseline transforms + show all edges.
    data.domains.forEach(d => applyDomainTransform(d.id));
    applyEdgeOffsetsFor(null);
  }

  function selectDomain(id) {
    clearSearchHighlight();
    if (soloFocusId !== id) enterSolo(id);
    document.querySelectorAll('.domain-bubble').forEach(b =>
      b.classList.toggle('active', b.dataset.id === id));
    document.querySelectorAll('.satellites-wrap').forEach(g => {
      const on = g.dataset.domain === id;
      g.style.pointerEvents = on ? 'auto' : 'none';
      g.querySelectorAll('.table-bubble').forEach((tb, i) => {
        if (on) setTimeout(() => tb.classList.add('visible'), i * 35);
        else tb.classList.remove('visible');
      });
    });
    document.querySelectorAll('.edge').forEach(l => {
      l.classList.toggle('hi', l.dataset.from === id || l.dataset.to === id);
    });
    renderSide(id);
  }

  function focusTable(fqn) {
    // Scroll the side panel to that table's card
    const card = document.querySelector(`[data-fqn-card="${CSS.escape(fqn)}"]`);
    if (card) card.scrollIntoView({behavior:'smooth', block:'center'});
  }

  function renderSide(domainId) {
    const d = domById[domainId];
    const tables = tablesByDom[domainId] || [];
    document.getElementById('welcome').hidden = true;
    const v = document.getElementById('domain-view');
    v.hidden = false;
    const bucketChip = b =>
      `<span class="chip bucket"><a href="../index.html#b-${b}">${b}</a></span>`;
    const layerChip = l => `<span class="chip layer-${l}">${l}</span>`;
    v.innerHTML = `
      <h2>${d.icon||''} ${escapeHtml(d.name)}</h2>
      <p class="desc">${escapeHtml(d.description)}</p>
      ${tables.length ? tables.map(t => `
        <div class="table-card ${t.layer} ${t.documented?'':'undocumented'}" data-fqn-card="${escapeAttr(t.fqn)}">
          <h3>${escapeHtml(t.short_name)} ${layerChip(t.layer)}</h3>
          <code class="fqn">${escapeHtml(t.fqn)}</code>
          <div class="body">${escapeHtml(t.description)}</div>
          <div class="meta">
            ${t.grain ? '<b>Grain:</b> '+escapeHtml(t.grain) : ''}
            ${t.partition_col ? ' · <b>Partition:</b> <code>'+escapeHtml(t.partition_col)+'</code>' : ''}
            ${t.source_system ? ' · <b>Source:</b> '+escapeHtml(t.source_system) : ''}
          </div>
          ${t.written_by.length || t.read_by.length ? `<div class="meta" style="margin-top:8px">
            ${t.written_by.length ? '<b>Written by:</b> '+t.written_by.map(bucketChip).join('') : ''}
            ${t.read_by.length ? (t.written_by.length?'<br>':'')+'<b>Read by:</b> '+t.read_by.map(bucketChip).join('') : ''}
          </div>` : ''}
          ${t.columns.length ? `<details><summary>📖 Data dictionary (${t.columns.length} columns)</summary>
            <table class="cols">
              <colgroup><col style="width:32%"><col style="width:18%"><col style="width:50%"></colgroup>
              <thead><tr><th>Column</th><th>Type</th><th>Description</th></tr></thead>
              <tbody>${t.columns.map(c => `
                <tr data-col-key="${escapeAttr(t.fqn + '::' + c.column)}">
                  <td><code>${escapeHtml(c.column)}</code>${c.pii?' <span class="pii" title="PII">🔒</span>':''}</td>
                  <td>${escapeHtml(c.type)}</td>
                  <td>${escapeHtml(c.description)}${c.domain_values?'<br><span class="meta">Values: '+escapeHtml(c.domain_values)+'</span>':''}</td>
                </tr>`).join('')}</tbody>
            </table>
          </details>` : `<div class="nodoc">📝 Data dictionary not yet documented.</div>`}
        </div>
      `).join('') : '<p class="empty">No tables in this domain yet.</p>'}
    `;
  }

  function escapeHtml(s) {
    return String(s ?? '').replace(/[&<>"']/g, c => ({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[c]));
  }
  function escapeAttr(s) { return escapeHtml(s); }

  svg.addEventListener('click', () => {
    if (suppressNextClick) { suppressNextClick = false; return; }
    clearSearchHighlight();
    document.querySelectorAll('.domain-bubble').forEach(b => b.classList.remove('active'));
    document.querySelectorAll('.satellites-wrap').forEach(g => {
      g.style.pointerEvents = 'none';
      g.querySelectorAll('.table-bubble').forEach(tb => tb.classList.remove('visible'));
    });
    document.querySelectorAll('.edge').forEach(l => l.classList.remove('hi'));
    document.getElementById('welcome').hidden = false;
    document.getElementById('domain-view').hidden = true;
    exitSolo();
  });

  document.addEventListener('keydown', e => { if (e.key === 'Escape') svg.dispatchEvent(new Event('click')); });
  const soloBackBtn = document.getElementById('solo-back');
  if (soloBackBtn) soloBackBtn.addEventListener('click', (e) => {
    e.stopPropagation();
    svg.dispatchEvent(new Event('click'));
  });
  // Expose for headless smoke tests.
  window.__wiki = { enterSolo, exitSolo };

  /* ─── Satellite label wrap helper ───────────────────────────────────── */
  /* Two-line wrap split on the underscore nearest to a 50/50 length split;
     each line clipped to MAX_CHARS with an ellipsis. The full name lives in
     the sibling <title> element (native browser tooltip). */
  function appendWrappedLabel(parent, name) {
    const MAX_CHARS = 17;
    const clip = s => s.length > MAX_CHARS ? s.slice(0, MAX_CHARS - 1) + '…' : s;
    if (name.length <= 14) {
      const t = el('text', {x:0, y:3}); t.textContent = name;
      parent.appendChild(t); return;
    }
    // Minimax split on '_': pick the underscore that minimizes max(left, right).
    let bestIdx = -1, bestMax = Infinity;
    for (let i = 0; i < name.length; i++) {
      if (name[i] !== '_') continue;
      const m = Math.max(i, name.length - i - 1);
      if (m < bestMax) { bestMax = m; bestIdx = i; }
    }
    let line1, line2;
    if (bestIdx >= 0) {
      line1 = name.slice(0, bestIdx);
      line2 = name.slice(bestIdx + 1);
    } else {
      const m = Math.ceil(name.length / 2);
      line1 = name.slice(0, m); line2 = name.slice(m);
    }
    const t = el('text', {x:0, y:0});
    const ts1 = el('tspan', {x:0, dy:'-0.1em'}); ts1.textContent = clip(line1);
    const ts2 = el('tspan', {x:0, dy:'1.15em'}); ts2.textContent = clip(line2);
    t.appendChild(ts1); t.appendChild(ts2);
    parent.appendChild(t);
  }

  /* ─── Spotlight-style search omnibox ────────────────────────────────── */
  /* Index: domains + table short_names + table fqns + columns. ~260 items.
     Match: case-insensitive substring. Rank: exact > prefix > word-boundary
     > substring; tiebreak by shorter-then-alpha. Results grouped by entity
     type, capped 5 per group. */
  const searchInput = document.getElementById('search');
  const searchBox = document.getElementById('search-results');
  const PER_GROUP_CAP = 5;

  // Build the searchable index once.
  const idx = { domain: [], table: [], column: [] };
  for (const d of data.domains) {
    idx.domain.push({ kind: 'domain', label: d.name, meta: d.description || '', id: d.id });
  }
  for (const t of data.tables) {
    if (!t.render) continue;
    // Table is searchable by both short_name and fqn (two index entries pointing to same target).
    idx.table.push({ kind: 'table', label: t.short_name, meta: t.fqn, fqn: t.fqn, domainId: t.domain_id });
    idx.table.push({ kind: 'table', label: t.fqn, meta: t.short_name, fqn: t.fqn, domainId: t.domain_id, _aliasFqn: true });
    for (const c of (t.columns || [])) {
      idx.column.push({ kind: 'column', label: c.column, meta: t.short_name, fqn: t.fqn, domainId: t.domain_id });
    }
  }

  function score(label, q) {
    // Lower is better. Negative = no match.
    const L = label.toLowerCase(), Q = q.toLowerCase();
    if (L === Q) return 0;
    if (L.startsWith(Q)) return 1;
    // Word-boundary: query starts at position right after a non-alnum char.
    const wb = new RegExp('(^|[^a-z0-9])' + Q.replace(/[.*+?^${}()|[\]\\]/g, '\\$&'));
    if (wb.test(L)) return 2;
    if (L.includes(Q)) return 3;
    return -1;
  }

  function rankGroup(items, q, opts={}) {
    const dedup = new Set();
    const scored = [];
    for (const it of items) {
      const s = score(it.label, q);
      if (s < 0) continue;
      // Collapse duplicate table entries (short_name + fqn alias point to same fqn) — keep best score.
      const dedupKey = opts.dedupBy ? it[opts.dedupBy] : null;
      if (dedupKey) {
        if (dedup.has(dedupKey)) continue;
        // Look ahead: prefer the non-alias entry if it scores equal-or-better.
      }
      scored.push({ it, s });
    }
    scored.sort((a, b) => {
      if (a.s !== b.s) return a.s - b.s;
      if (a.it.label.length !== b.it.label.length) return a.it.label.length - b.it.label.length;
      return a.it.label.localeCompare(b.it.label);
    });
    const out = [];
    for (const e of scored) {
      const k = opts.dedupBy ? e.it[opts.dedupBy] : null;
      if (k) { if (dedup.has(k)) continue; dedup.add(k); }
      out.push(e.it);
    }
    return out;
  }

  function highlight(label, q) {
    const i = label.toLowerCase().indexOf(q.toLowerCase());
    if (i < 0) return escapeHtml(label);
    return escapeHtml(label.slice(0, i)) + '<b>' + escapeHtml(label.slice(i, i + q.length)) + '</b>' + escapeHtml(label.slice(i + q.length));
  }

  let activeRowIdx = -1, flatRows = [];

  function renderResults(q) {
    flatRows = []; activeRowIdx = -1;
    if (!q || q.trim().length < 1) { closeResults(); return; }
    q = q.trim();

    const groups = [
      { key: 'domain',  head: 'Domains', hits: rankGroup(idx.domain, q) },
      { key: 'table',   head: 'Tables',  hits: rankGroup(idx.table, q, {dedupBy: 'fqn'}) },
      { key: 'column',  head: 'Columns', hits: rankGroup(idx.column, q) },
    ];

    const totalHits = groups.reduce((a, g) => a + g.hits.length, 0);
    if (totalHits === 0) {
      searchBox.innerHTML = '<div class="sr-empty">No matches for <code>' + escapeHtml(q) + '</code></div>';
      openResults(); return;
    }

    let html = '';
    for (const g of groups) {
      if (!g.hits.length) continue;
      const shown = g.hits.slice(0, PER_GROUP_CAP);
      const overflow = g.hits.length - shown.length;
      html += '<div class="sr-group"><div class="sr-head">' + g.head + ' · ' + g.hits.length + '</div>';
      for (const it of shown) {
        const ridx = flatRows.length;
        flatRows.push(it);
        const meta = it.kind === 'column' ? '→ ' + escapeHtml(it.meta)
                   : it.kind === 'table'  ? escapeHtml(it.meta)
                   : escapeHtml(it.meta || '').slice(0, 60);
        html += '<div class="sr-row" data-ridx="' + ridx + '" role="option">' +
                '<span class="sr-name">' + highlight(it.label, q) + '</span>' +
                '<span class="sr-meta">' + meta + '</span></div>';
      }
      if (overflow > 0) {
        html += '<div class="sr-more">…and ' + overflow + ' more — refine your query.</div>';
      }
      html += '</div>';
    }
    searchBox.innerHTML = html;
    openResults();
    activeRowIdx = 0;
    paintActive();
  }

  function openResults() {
    searchBox.classList.add('open');
    searchInput.setAttribute('aria-expanded', 'true');
  }
  function closeResults() {
    searchBox.classList.remove('open');
    searchInput.setAttribute('aria-expanded', 'false');
    activeRowIdx = -1; flatRows = [];
  }
  function paintActive() {
    searchBox.querySelectorAll('.sr-row').forEach(r => r.classList.remove('active'));
    if (activeRowIdx >= 0) {
      const r = searchBox.querySelector('.sr-row[data-ridx="' + activeRowIdx + '"]');
      if (r) { r.classList.add('active'); r.scrollIntoView({block:'nearest'}); }
    }
  }

  function activate(it) {
    closeResults();
    searchInput.value = '';
    searchInput.blur();
    const inTableMode = document.body.classList.contains('view-table');
    if (it.kind === 'domain') {
      if (inTableMode) {
        const sec = document.querySelector('#tm-inner .tm-domain[data-domain="' + cssEscape(it.id) + '"]');
        if (sec) sec.scrollIntoView({behavior:'smooth', block:'start'});
      }
      selectDomain(it.id);
      return;
    }
    // Table or column: select parent domain, then scroll/expand.
    if (inTableMode) {
      highlightTableModelCard(it.fqn);
      if (it.kind === 'column') {
        requestAnimationFrame(() => {
          const c = document.querySelector('#tm-inner [data-tm-fqn="' + cssEscape(it.fqn) + '"]');
          const det = c && c.querySelector('details');
          if (det) det.open = true;
          const row = c && c.querySelector('tr[data-tm-col="' + cssEscape(it.fqn + '::' + it.label) + '"]');
          if (row) {
            row.scrollIntoView({behavior:'smooth', block:'center'});
            row.classList.remove('col-flash'); void row.offsetWidth; row.classList.add('col-flash');
            setTimeout(() => row.classList.remove('col-flash'), 1700);
          }
        });
      }
      return;
    }
    selectDomain(it.domainId);
    // Defer one frame so the side panel is rendered before we look for cards.
    requestAnimationFrame(() => {
      // Highlight the matching satellite bubble (and dim siblings) so the eye
      // lands on the right table when 15+ satellites ring a single domain.
      highlightSatelliteByFqn(it.fqn);
      const card = document.querySelector('[data-fqn-card="' + cssEscape(it.fqn) + '"]');
      if (!card) return;
      card.scrollIntoView({behavior:'smooth', block:'center'});
      if (it.kind === 'column') {
        const det = card.querySelector('details');
        if (det) det.open = true;
        // Wait for details to expand, then scroll & flash the row.
        requestAnimationFrame(() => {
          const row = card.querySelector('tr[data-col-key="' + cssEscape(it.fqn + '::' + it.label) + '"]');
          if (!row) return;
          row.scrollIntoView({behavior:'smooth', block:'center'});
          row.classList.remove('col-flash');
          // Force reflow so the animation restarts even if class was just present.
          void row.offsetWidth;
          row.classList.add('col-flash');
          setTimeout(() => row.classList.remove('col-flash'), 1700);
        });
      }
    });
  }

  function cssEscape(s) {
    return (window.CSS && CSS.escape) ? CSS.escape(s) : String(s).replace(/["\\]/g, '\\$&');
  }

  // Input handlers
  let lastQ = '';
  searchInput.addEventListener('input', () => {
    const q = searchInput.value;
    if (q === lastQ) return;
    lastQ = q;
    renderResults(q);
  });
  searchInput.addEventListener('focus', () => {
    if (searchInput.value.trim()) renderResults(searchInput.value);
  });
  searchInput.addEventListener('keydown', e => {
    if (e.key === 'ArrowDown') {
      if (!flatRows.length) return;
      e.preventDefault();
      activeRowIdx = (activeRowIdx + 1) % flatRows.length;
      paintActive();
    } else if (e.key === 'ArrowUp') {
      if (!flatRows.length) return;
      e.preventDefault();
      activeRowIdx = (activeRowIdx - 1 + flatRows.length) % flatRows.length;
      paintActive();
    } else if (e.key === 'Enter') {
      if (activeRowIdx >= 0 && flatRows[activeRowIdx]) {
        e.preventDefault();
        activate(flatRows[activeRowIdx]);
      }
    } else if (e.key === 'Escape') {
      e.preventDefault();
      if (searchBox.classList.contains('open')) closeResults();
      else { searchInput.value = ''; lastQ = ''; searchInput.blur(); }
      e.stopPropagation();
    }
  });
  searchBox.addEventListener('mousedown', e => {
    const row = e.target.closest('.sr-row');
    if (!row) return;
    e.preventDefault();  // keep input focused until activate() blurs it
    const r = flatRows[+row.dataset.ridx];
    if (r) activate(r);
  });
  document.addEventListener('click', e => {
    if (!e.target.closest('.search-wrap')) closeResults();
  });

  // Global keyboard: '/' and Ctrl/Cmd+K focus the omnibox.
  document.addEventListener('keydown', e => {
    const inField = /^(INPUT|TEXTAREA|SELECT)$/.test(document.activeElement.tagName);
    if (!inField && e.key === '/') {
      e.preventDefault(); searchInput.focus(); searchInput.select();
    } else if ((e.ctrlKey || e.metaKey) && e.key.toLowerCase() === 'k') {
      e.preventDefault(); searchInput.focus(); searchInput.select();
    }
  });
})();
</script>
</body>
</html>
"""


# ─── Entry point ───────────────────────────────────────────────────────────

def _is_stale() -> bool:
    """Used by rebuild_wiki.py's mtime gate. True if `semantic.json` (the
    authoring SoT) is newer than the published HTML (i.e. a rebuild is
    needed)."""
    if not JSON_PATH.exists():
        return False
    if not OUT_HTML.exists():
        return True
    return JSON_PATH.stat().st_mtime > OUT_HTML.stat().st_mtime


def _write_outputs(html: str) -> None:
    """Write the rendered HTML to <ROOT>/semantic/.

    Per ADR 0005, the xlsx is the snapshotted artifact (see `_backup_xlsx`,
    called by `main()` BEFORE rendering). This function only writes the HTML
    itself — no HTML backup is kept.

    All writes are confined to the current tree's ROOT (handovers/ when run
    from handovers/scripts/, handovers/source/ when run from the replica).
    No mirroring or cross-tree copying is performed.
    """
    OUT_HTML.parent.mkdir(parents=True, exist_ok=True)
    OUT_HTML.write_text(html, encoding="utf-8")
    print(f"  wrote    {OUT_HTML.relative_to(ROOT.parent)}  ({len(html):,} bytes)")


def main(argv: list[str] | None = None) -> int:
    p = argparse.ArgumentParser(description="Build the semantic mini-wiki page.")
    p.add_argument("--seed", action="store_true",
                   help="Create a starter semantic.json (refuses to overwrite).")
    p.add_argument("--import-xlsx", action="store_true",
                   help="Read semantic.xlsx, validate, write semantic.json. "
                        "Use for first bootstrap when only the xlsx exists.")
    p.add_argument("--export-xlsx", action="store_true",
                   help="Read semantic.json, snapshot the current xlsx to backup/, "
                        "write a fresh semantic.xlsx. Use after editor sessions.")
    p.add_argument("--check-stale", action="store_true",
                   help="Exit 0 if HTML is up-to-date, 1 if rebuild is needed. No output.")
    args = p.parse_args(argv)

    if args.check_stale:
        return 1 if _is_stale() else 0

    if args.seed:
        _seed()
        return 0

    if getattr(args, "import_xlsx", False):
        if not XLSX.exists():
            print(f"✗ {XLSX} does not exist — nothing to import.", file=sys.stderr)
            return 2
        print(f"  reading  {XLSX.relative_to(ROOT.parent)}")
        data = _read_xlsx()
        rpt = BuildReport()
        _validate(data, rpt)
        if not rpt.ok:
            rpt.print_summary()
            print("\n✗ Import aborted — fix the errors above and re-run.", file=sys.stderr)
            return 1
        _write_json(data)
        print(f"  wrote    {JSON_PATH.relative_to(ROOT.parent)}")
        rpt.print_summary()
        return 0

    if getattr(args, "export_xlsx", False):
        if not JSON_PATH.exists():
            print(f"✗ {JSON_PATH} does not exist — nothing to export.", file=sys.stderr)
            return 2
        print(f"  reading  {JSON_PATH.relative_to(ROOT.parent)}")
        data = _read_json()
        bkp = _backup_xlsx()
        if bkp:
            print(f"  backup   {bkp}")
        _write_xlsx(data)
        print(f"  wrote    {XLSX.relative_to(ROOT.parent)}  ({XLSX.stat().st_size:,} bytes)")
        return 0

    # Default = build from JSON.
    if not JSON_PATH.exists():
        if XLSX.exists():
            print(f"✗ {JSON_PATH} does not exist. Bootstrap with:", file=sys.stderr)
            print(f"    python handovers/scripts/build_semantic.py --import-xlsx", file=sys.stderr)
        else:
            print(f"✗ {JSON_PATH} does not exist. Run with --seed first.", file=sys.stderr)
        return 2

    print(f"  reading  {JSON_PATH.relative_to(ROOT.parent)}")
    data = _read_json()
    rpt = BuildReport()
    _validate(data, rpt)

    if not rpt.ok:
        rpt.print_summary()
        print("\n✗ Build aborted — fix the errors above and re-run.", file=sys.stderr)
        return 1

    # Task-1: snapshot the human-readable xlsx (if any) before each build.
    bkp = _backup_xlsx()
    if bkp:
        print(f"  backup   {bkp}")

    html = _render(data)
    _write_outputs(html)
    rpt.print_summary()
    return 0


if __name__ == "__main__":
    sys.exit(main())
