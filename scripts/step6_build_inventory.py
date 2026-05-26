"""
Build handovers/inventory.xlsx — a 3-sheet template + report.

Sheets:
  00_README   — column dictionary, allowed values, ID regex, import notes
  01_Buckets  — master bucket list, every column maps 1:1 to an admin form field
  02_Manuals  — manuals reading list, every column maps 1:1 to an admin form field

This file is BOTH a generated report (for Excel-only reviewers) AND the
import template consumed by admin.html → POST /upload-inventory (PR-C).
Keeping one file with one schema avoids template-drift bugs.

⚠ Editing the xlsx directly and saving does nothing UNLESS you upload it
back through admin.html's "Import" button. The single source of truth on
disk remains `handovers/inventory.json`.
"""
from __future__ import annotations
import json
import re
import sys
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment

ROOT     = Path(__file__).resolve().parents[1]   # handovers/
INV_JSON = ROOT / "inventory.json"
OUT      = ROOT / "inventory.xlsx"
# Old name from before May-2026 — removed on every rebuild so stale copies
# don't get edited and lost.
OLD_OUT  = ROOT / "handover_inventory.xlsx"

# ─── Schema ────────────────────────────────────────────────────────────────
# (excel_header, json_key, width, comment)
BUCKET_COLS: list[tuple[str, str, int, str]] = [
    ("Bucket ID",   "bucket_id",   14, "Required. Pattern: CAT-YYYY-NN (e.g. PRJ-2026-01). CAT ∈ {PRJ,CMP,MOD,BAU,STR,ADH}."),
    ("Name",        "name",        45, "Required. Short human-readable label."),
    ("Category",    "category",    11, "Required. One of: Project, Campaign, Model, BAU, Strategy, Adhoc."),
    ("Status",      "status",      16, "One of: Active, Completed, Superseded, On-hold, Retired."),
    ("Tier",        "tier",         6, "One of: P0, P1, P2."),
    ("Year",        "year",         7, "Integer (e.g. 2026)."),
    ("Source plan", "source_plan", 28, "Free text — provenance note (e.g. 'scan-usecases-2026.md')."),
    ("Repo link",   "repo_link",   50, "Comma-separated repo paths (e.g. 'marketing/foo/, sales/bar/')."),
    ("Repo role",   "repo_role",   28, "Free text. Used for §07 Repo-map labels."),
    ("Lineage",     "lineage",     35, "Free text — upstream dependencies / data lineage notes."),
    ("Purpose",     "purpose",     60, "Required. One-sentence purpose (becomes card description)."),
]

MANUAL_COLS: list[tuple[str, str, int, str]] = [
    ("Manual ID", "id",    14, "Required. Pattern: MAN-YYYY-NN (e.g. MAN-2026-01)."),
    ("Title",     "title", 45, "Required. Manual title shown on the card."),
    ("Desc",      "desc",  60, "One-sentence description."),
    ("File",      "file",  45, "Filename (resolved against wiki root by default). Leave blank if URL is set."),
    ("URL",       "url",   70, "Optional. Direct URL (http(s) / SharePoint). Takes precedence over File."),
    ("Kind",      "kind",  10, "Optional. One of: doc, pdf, xlsx, pptx, video, html. Choose '(auto)' or leave blank to infer from file extension."),
]

# ─── Style helpers ────────────────────────────────────────────────────────
HDR_FILL  = PatternFill("solid", fgColor="1F4E78")
HDR_FONT  = Font(bold=True, color="FFFFFF", size=11)
HDR_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
WRAP      = Alignment(vertical="top", wrap_text=True)
BORDER    = Border(*(Side(style="thin", color="CCCCCC"),) * 4)


def _load() -> dict:
    if not INV_JSON.exists():
        sys.exit(f"ERROR: {INV_JSON} not found. Run migrate_to_json.py once.")
    return json.loads(INV_JSON.read_text(encoding="utf-8"))


def _write_data_sheet(wb: Workbook, name: str, cols: list[tuple[str, str, int, str]],
                       items: list[dict], skip_retired: bool = True) -> None:
    """Write one data sheet with headers, comments, table styling, and
    drop-down validations for known-vocabulary columns."""
    ws = wb.create_sheet(name)
    headers = [c[0] for c in cols]
    keys    = [c[1] for c in cols]
    widths  = [c[2] for c in cols]
    notes   = [c[3] for c in cols]

    ws.append(headers)
    for i, cell in enumerate(ws[1]):
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = HDR_ALIGN
        if notes[i]:
            c = Comment(notes[i], "schema")
            c.width = 360
            c.height = 80
            cell.comment = c

    rows_written = 0
    for r in items:
        if skip_retired and str(r.get("status", "")).startswith("Retired"):
            continue
        ws.append([r.get(k, "") for k in keys])
        rows_written += 1

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = WRAP
            cell.border = BORDER

    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    # data validations for vocab columns
    def _add_dv(col_idx: int, formula: str, prompt: str) -> None:
        dv = DataValidation(type="list", formula1=formula, allow_blank=True,
                            showDropDown=False, errorStyle="warning")
        dv.prompt = prompt
        dv.promptTitle = "Allowed values"
        dv.error = "Value not in list — will be flagged on import."
        dv.errorTitle = "Unknown value"
        col_letter = get_column_letter(col_idx)
        dv.add(f"{col_letter}2:{col_letter}10000")
        ws.add_data_validation(dv)

    if "category" in keys:
        _add_dv(keys.index("category") + 1,
                '"Project,Campaign,Model,BAU,Strategy,Adhoc"',
                "Category bucket.")
    if "status" in keys:
        _add_dv(keys.index("status") + 1,
                '"Active,Completed,Superseded,On-hold,Retired"',
                "Lifecycle status.")
    if "tier" in keys:
        _add_dv(keys.index("tier") + 1, '"P0,P1,P2"', "Priority tier.")
    if "kind" in keys:
        _add_dv(keys.index("kind") + 1,
                '"(auto),doc,pdf,xlsx,pptx,video,html"',
                "Document kind. Choose '(auto)' (or leave blank) to auto-infer from the file extension.")

    if rows_written:
        last_col = get_column_letter(len(headers))
        last_row = rows_written + 1
        safe_name = "tbl_" + re.sub(r"\W", "_", name)
        tbl = Table(displayName=safe_name, ref=f"A1:{last_col}{last_row}")
        tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws.add_table(tbl)


# ─── README sheet ────────────────────────────────────────────────────────
def _readme(data: dict) -> list[list]:
    n_b = sum(1 for b in data.get("buckets", []) if not str(b.get("status", "")).startswith("Retired"))
    n_m = len(data.get("manuals", []))
    return [
        ["VN Analytics Handover Inventory", "", ""],
        ["Source",    "handovers/inventory.json", "single source of truth (edit via admin.html)"],
        ["Generated", datetime.now().strftime("%Y-%m-%d %H:%M"), "auto-built by step6_build_inventory.py"],
        ["", "", ""],
        ["Sheet", "Rows", "Purpose"],
        ["00_README",  "—",       "This page — schema + import notes."],
        ["01_Buckets", str(n_b),  "Master bucket list (all 6 categories)."],
        ["02_Manuals", str(n_m),  "Team-document reading list."],
        ["", "", ""],
        ["How to import this file back into admin.html", "", ""],
        ["1.", "Open admin.html (run `python handovers/serve_admin.py`).", ""],
        ["2.", "Click ⬆ Import (header toolbar) and choose this .xlsx file.", ""],
        ["3.", "Review the per-row diff: keep mine / take theirs / skip.", ""],
        ["4.", "Click Apply — admin will PUT a new inventory.json (.bak rotated automatically).", ""],
        ["", "", ""],
        ["Validation rules (enforced on import)", "", ""],
        ["Bucket ID", "must match regex ^[A-Z]{3}-\\d{4}-\\d{2}$",
         "e.g. PRJ-2026-01. Prefix must match category (PRJ→Project, CMP→Campaign, MOD→Model, BAU→BAU, STR→Strategy, ADH→Adhoc)."],
        ["Manual ID", "must match regex ^MAN-\\d{4}-\\d{2}$", "e.g. MAN-2026-01."],
        ["Category",  "one of: Project, Campaign, Model, BAU, Strategy, Adhoc", "drop-down enforced"],
        ["Status",    "one of: Active, Completed, Superseded, On-hold, Retired", "drop-down enforced"],
        ["Tier",      "one of: P0, P1, P2", "drop-down enforced"],
        ["Kind",      "one of: doc, pdf, xlsx, pptx, video", "auto-inferred from URL/file extension if blank"],
    ]


def main() -> None:
    data = _load()
    buckets = data.get("buckets", [])
    manuals = data.get("manuals", [])

    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("00_README")
    for row in _readme(data):
        ws.append(row)
    ws["A1"].font = Font(bold=True, size=14, color="1F4E78")
    for cell in ws[5]:
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
    for r_idx in (10, 16):
        for c in ws[r_idx]:
            c.font = Font(bold=True, color="1F4E78")
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 70
    for r in ws.iter_rows():
        for c in r:
            c.alignment = Alignment(vertical="top", wrap_text=True)

    _write_data_sheet(wb, "01_Buckets", BUCKET_COLS, buckets)
    _write_data_sheet(wb, "02_Manuals", MANUAL_COLS, manuals, skip_retired=False)

    wb.save(OUT)
    if OLD_OUT.exists():
        try:
            OLD_OUT.unlink()
            print(f"Removed legacy {OLD_OUT.name} (renamed to {OUT.name})")
        except OSError:
            pass
    print(f"Wrote {OUT} — {len(wb.sheetnames)} sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
