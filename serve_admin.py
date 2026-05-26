"""
serve_admin.py
==============

Tiny stdlib-only static server for the local handovers tooling.
Binds to all interfaces (0.0.0.0) so the wiki is reachable from the LAN
(e.g. http://10.251.132.71:8765/) — useful for sharing read-only views
with teammates on the same network. Admin / semantic-editor pages and
all write endpoints are gated by a shared 6-digit passcode (default
"111111"; changeable from admin.html).

CLI:
  python serve_admin.py                  # → opens admin.html (default)
  python serve_admin.py --semantic       # → opens semantic_editor.html
  python serve_admin.py --both           # → opens both
  python serve_admin.py --no-browser     # → server only, no auto-open
  python serve_admin.py --open index.html  # → opens any page
  python serve_admin.py --port 9000      # → bind to a different port

Env vars (back-compat):
  SERVE_ADMIN_NO_BROWSER=1   suppress auto-open
  SERVE_ADMIN_PAGE=...       override default page (admin.html / BOTH / etc.)

Auth:
  - The first run creates `.passcode` (JSON: {hash, salt}) with the default
    pin "111111". The file is gitignored and never logged.
  - On successful POST /api/auth/login the server writes a random bearer
    token to `.session`; the browser stores it in localStorage and sends
    it as `Authorization: Bearer <token>` on every same-origin request.
  - Public GETs (index.html, wiki pages, *.json reads) are NOT gated, so
    teammates can still browse the wiki over the LAN. Only admin.html,
    semantic_editor.html and write endpoints require the token.

Endpoints:
  GET  /<anything>             → static file (admin.html, inventory.json,
                                  semantic_editor.html, semantic.json, …)
  POST /api/auth/login         → {passcode} → {token}
  POST /api/auth/whoami        → bearer → 200/401
  POST /api/auth/change        → bearer + {current,new} → {token}
  PUT  /inventory.json         → overwrite inventory.json (rotates .bak)   [auth]
  PUT  /semantic.json          → overwrite semantic.json (rotates .bak,    [auth]
                                  then rebuilds 08_Semantic_Model.html inline)
  POST /upload-inventory       → parse an xlsx body, return per-row diff JSON
                                  (client applies + saves via the standard PUT) [auth]
  POST /upload-semantic        → parse semantic.xlsx body, write semantic.json
                                  directly (one-shot bootstrap / re-import).
                                  Backs up existing semantic.json to .bak.  [auth]
  POST /export-semantic-xlsx   → snapshot current semantic.xlsx into backup/,
                                  then write a fresh semantic.xlsx from
                                  semantic.json. No request body required.  [auth]

Usage (from any directory):
    python ./serve_admin.py

Press Ctrl-C to stop.
"""
from __future__ import annotations
import http.server
import socketserver
import webbrowser
import threading
import os
import sys
import json
import re
import io
import hashlib
import secrets
import socket
from functools import partial
from pathlib import Path

ROOT = Path(__file__).resolve().parent          # handovers/
PORT = 8765
URL  = f"http://localhost:{PORT}/admin.html"

# Make `import build_semantic` work — scripts/ is not a package.
_SCRIPTS_DIR = ROOT / "scripts"
if str(_SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(_SCRIPTS_DIR))

# Files the editor pages are allowed to write via HTTP PUT (relative to ROOT).
# Anything else returns 403 — prevents accidental writes to source files.
_WRITABLE_FILES = {"inventory.json", "semantic/semantic.json"}

# ── Passcode + session helpers ──────────────────────────────────────────
_PASSCODE_FILE = ROOT / ".passcode"
_SESSION_FILE  = ROOT / ".session"
_DEFAULT_PIN   = "111111"
_PIN_RX        = re.compile(r"^\d{6}$")
# Pages that require a valid bearer token to load. Everything else
# (index.html, the wiki, *.json reads) stays public.
_GATED_PAGES   = {"admin.html", "semantic_editor.html"}
# Write endpoints that require a valid bearer token.
_AUTH_POSTS    = {"upload-inventory", "upload-semantic", "export-semantic-xlsx"}
_AUTH_API      = {"api/auth/login", "api/auth/whoami", "api/auth/change"}


def _hash_pin(pin: str, salt: str) -> str:
    """SHA-256(salt|pin) — adequate for a LAN shared-secret; not a password hash."""
    return hashlib.sha256(f"{salt}|{pin}".encode("utf-8")).hexdigest()


def _load_passcode() -> dict:
    """Read .passcode; bootstrap it to the default pin on first run."""
    if _PASSCODE_FILE.exists():
        try:
            return json.loads(_PASSCODE_FILE.read_text(encoding="utf-8"))
        except Exception:
            pass  # fall through to re-bootstrap
    salt = secrets.token_hex(8)
    rec = {"hash": _hash_pin(_DEFAULT_PIN, salt), "salt": salt}
    _PASSCODE_FILE.write_text(json.dumps(rec), encoding="utf-8")
    return rec


def _verify_pin(pin: str) -> bool:
    if not _PIN_RX.match(pin or ""):
        return False
    rec = _load_passcode()
    return _hash_pin(pin, rec["salt"]) == rec["hash"]


def _set_pin(new_pin: str) -> None:
    salt = secrets.token_hex(8)
    _PASSCODE_FILE.write_text(
        json.dumps({"hash": _hash_pin(new_pin, salt), "salt": salt}),
        encoding="utf-8",
    )


def _issue_token() -> str:
    tok = secrets.token_hex(24)
    _SESSION_FILE.write_text(tok, encoding="utf-8")
    return tok


def _current_token() -> str | None:
    if not _SESSION_FILE.exists():
        return None
    try:
        v = _SESSION_FILE.read_text(encoding="utf-8").strip()
        return v or None
    except Exception:
        return None


def _bearer_from_headers(headers) -> str | None:
    auth = headers.get("Authorization") or ""
    m = re.match(r"Bearer\s+([A-Fa-f0-9]+)", auth)
    return m.group(1) if m else None


def _is_authed(headers) -> bool:
    tok_in = _bearer_from_headers(headers)
    tok_disk = _current_token()
    return bool(tok_in and tok_disk and secrets.compare_digest(tok_in, tok_disk))


def _lan_ips() -> list[str]:
    """Best-effort list of non-loopback IPv4 addresses for the host."""
    out = set()
    try:
        host = socket.gethostname()
        for fam, _t, _p, _c, sockaddr in socket.getaddrinfo(host, None):
            if fam == socket.AF_INET:
                ip = sockaddr[0]
                if ip and not ip.startswith("127."):
                    out.add(ip)
    except Exception:
        pass
    # UDP-connect trick — finds the IP used to reach the default gateway.
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        try:
            s.connect(("8.8.8.8", 80))
            ip = s.getsockname()[0]
            if ip and not ip.startswith("127."):
                out.add(ip)
        finally:
            s.close()
    except Exception:
        pass
    return sorted(out)

# ── xlsx import helpers ─────────────────────────────────────────────────
_BUCKET_HDR_MAP = {
    "bucket id": "bucket_id", "name": "name", "category": "category",
    "status": "status", "tier": "tier", "year": "year",
    "source plan": "source_plan", "repo link": "repo_link",
    "repo role": "repo_role", "lineage": "lineage", "purpose": "purpose",
}
_MANUAL_HDR_MAP = {
    "manual id": "id", "id": "id", "title": "title", "desc": "desc",
    "description": "desc", "file": "file", "url": "url", "kind": "kind",
}
_BUCKET_ID_RX = re.compile(r"^[A-Z]{3}-\d{4}-\d{2,3}$")
_MANUAL_ID_RX = re.compile(r"^MAN-\d{4}-\d{2,3}$")
_PREFIX_CAT   = {"PRJ":"Project","CMP":"Campaign","MOD":"Model",
                 "BAU":"BAU","STR":"Strategy","ADH":"Adhoc"}
_VALID_STATUS = {"Active","Completed","Superseded","On-hold","Retired"}
_VALID_TIER   = {"P0","P1","P2"}


def _read_sheet(ws, hdr_map: dict) -> tuple[list[dict], list[str]]:
    """Read an openpyxl worksheet → list of dicts (only known columns)
    + list of validation errors."""
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return [], [f"sheet '{ws.title}' is empty"]
    col_keys: list[str | None] = []
    for h in header_row:
        if h is None:
            col_keys.append(None)
            continue
        key = hdr_map.get(str(h).strip().lower())
        col_keys.append(key)
    if not any(col_keys):
        return [], [f"sheet '{ws.title}' header row has no recognised columns"]
    out: list[dict] = []
    errs: list[str] = []
    for r_idx, row in enumerate(rows_iter, start=2):
        rec: dict = {}
        for k, v in zip(col_keys, row):
            if not k:
                continue
            if v is None:
                rec[k] = ""
            elif isinstance(v, float) and v.is_integer():
                rec[k] = int(v)
            else:
                rec[k] = v if isinstance(v, (int, float)) else str(v).strip()
        # skip blank rows
        if not any(str(v).strip() for v in rec.values()):
            continue
        out.append(rec)
    return out, errs


def _validate_bucket(rec: dict) -> list[str]:
    e = []
    bid = str(rec.get("bucket_id") or "")
    cat = str(rec.get("category") or "")
    if not _BUCKET_ID_RX.match(bid):
        e.append(f"bucket_id '{bid}' must match ^[A-Z]{{3}}-\\d{{4}}-\\d{{2,3}}$")
    elif cat and _PREFIX_CAT.get(bid[:3]) != cat:
        e.append(f"bucket_id prefix '{bid[:3]}' does not match category '{cat}' "
                 f"(expected {_PREFIX_CAT.get(bid[:3]) or 'unknown'})")
    if not str(rec.get("name") or "").strip():
        e.append("name is required")
    if cat and cat not in _PREFIX_CAT.values():
        e.append(f"category '{cat}' not in {sorted(_PREFIX_CAT.values())}")
    st = str(rec.get("status") or "")
    if st and st not in _VALID_STATUS:
        e.append(f"status '{st}' not in {sorted(_VALID_STATUS)}")
    tr = str(rec.get("tier") or "")
    if tr and tr not in _VALID_TIER:
        e.append(f"tier '{tr}' not in {sorted(_VALID_TIER)}")
    return e


def _validate_manual(rec: dict) -> list[str]:
    e = []
    mid = str(rec.get("id") or "")
    if not _MANUAL_ID_RX.match(mid):
        e.append(f"id '{mid}' must match ^MAN-\\d{{4}}-\\d{{2,3}}$")
    if not str(rec.get("title") or "").strip():
        e.append("title is required")
    # `(auto)` is the dropdown placeholder for "infer kind from file extension"
    # — normalise to empty so downstream auto-detection (step11) takes over.
    if str(rec.get("kind") or "").strip().lower() in ("(auto)", "auto"):
        rec["kind"] = ""
    return e


def _diff_rows(existing: list[dict], incoming: list[dict],
                key: str, validate) -> dict:
    """Compare incoming list against existing by primary key.

    Returns:
        {
          additions: [rec, ...],          # not in existing
          conflicts: [{id, mine, theirs, changed_fields}, ...],
          unchanged: [id, ...],
          errors:    [{row, id, problems:[...]}, ...]
        }
    """
    by_id = {str(r.get(key) or ""): r for r in existing}
    additions, conflicts, unchanged, errors = [], [], [], []
    for i, rec in enumerate(incoming, start=2):
        problems = validate(rec)
        rid = str(rec.get(key) or "")
        if problems:
            errors.append({"row": i, "id": rid, "problems": problems})
            continue
        if rid not in by_id:
            additions.append(rec)
        else:
            mine = by_id[rid]
            changed = [k for k in rec.keys()
                       if str(mine.get(k, "") or "") != str(rec.get(k, "") or "")]
            if changed:
                conflicts.append({"id": rid, "mine": mine, "theirs": rec,
                                  "changed_fields": changed})
            else:
                unchanged.append(rid)
    return {"additions": additions, "conflicts": conflicts,
            "unchanged": unchanged, "errors": errors}


def _build_diff(xlsx_bytes: bytes) -> dict:
    """Parse an xlsx byte buffer; return diff vs current inventory.json."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise RuntimeError("openpyxl not installed — `pip install openpyxl`")
    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=True)
    out: dict = {"sheets_found": wb.sheetnames}

    # find sheets by name (case-insensitive, accept legacy names)
    def _find(name_substrings: list[str]):
        for n in wb.sheetnames:
            low = n.lower()
            if any(s in low for s in name_substrings):
                return wb[n]
        return None

    inv_path = ROOT / "inventory.json"
    current = json.loads(inv_path.read_text(encoding="utf-8")) if inv_path.exists() else {}

    bucket_ws = _find(["bucket"])
    manual_ws = _find(["manual"])
    if not bucket_ws and not manual_ws:
        raise RuntimeError("workbook has no 'Buckets' or 'Manuals' sheet")

    if bucket_ws:
        rows, _ = _read_sheet(bucket_ws, _BUCKET_HDR_MAP)
        out["buckets"] = _diff_rows(current.get("buckets", []), rows,
                                     "bucket_id", _validate_bucket)
    if manual_ws:
        rows, _ = _read_sheet(manual_ws, _MANUAL_HDR_MAP)
        out["manuals"] = _diff_rows(current.get("manuals", []), rows,
                                     "id", _validate_manual)
    return out


# ── HTTP handler ────────────────────────────────────────────────────────
class _Handler(http.server.SimpleHTTPRequestHandler):
    """Serve the handovers/ folder.
    PUT  /inventory.json     overwrites inventory.json (rotates .bak)
    POST /upload-inventory   parses raw xlsx body, returns JSON diff
    """

    def end_headers(self):
        # Disable caching so a saved JSON is always re-read on reload.
        self.send_header("Cache-Control", "no-store, must-revalidate")
        self.send_header("Pragma", "no-cache")
        super().end_headers()

    def _path_no_query(self) -> str:
        return self.path.lstrip("/").split("?", 1)[0]

    @staticmethod
    def _etag(data: bytes) -> str:
        import hashlib
        return '"' + hashlib.sha1(data).hexdigest() + '"'

    def do_GET(self):
        # Add an ETag header to JSON files we serve so the editor pages can do
        # optimistic-concurrency on PUT (If-Match).
        rel = self._path_no_query()
        if rel in _WRITABLE_FILES:
            target = (ROOT / rel)
            if target.exists():
                body = target.read_bytes()
                etag = self._etag(body)
                self.send_response(200)
                self.send_header("Content-Type", "application/json; charset=utf-8")
                self.send_header("Content-Length", str(len(body)))
                self.send_header("ETag", etag)
                self.end_headers()
                self.wfile.write(body)
                return
        super().do_GET()

    def do_PUT(self):
        # Allow editor pages to overwrite whitelisted JSON files without FSA.
        # Path is restricted to filenames inside ROOT to prevent traversal.
        try:
            if not _is_authed(self.headers):
                return self._send_401()
            rel = self._path_no_query()
            target = (ROOT / rel).resolve()
            if ROOT.resolve() not in target.parents and target != ROOT.resolve():
                self.send_error(403, "outside server root")
                return
            if rel not in _WRITABLE_FILES:
                self.send_error(403, f"only {sorted(_WRITABLE_FILES)} are writable")
                return
            # Optimistic-concurrency: if the client sent If-Match, verify it
            # matches the etag of the file currently on disk. This prevents a
            # second tab silently overwriting changes saved from a first tab.
            if_match = self.headers.get("If-Match")
            if if_match and target.exists():
                current_etag = self._etag(target.read_bytes())
                if if_match.strip() != current_etag:
                    self.send_response(412)
                    self.send_header("ETag", current_etag)
                    self.end_headers()
                    return
            length = int(self.headers.get("Content-Length", "0"))
            body = self.rfile.read(length)
            # rotate .bak
            if target.exists():
                bak = target.with_suffix(target.suffix + ".bak")
                bak.write_bytes(target.read_bytes())
            target.write_bytes(body)
            new_etag = self._etag(body)

            # Hook: after a semantic.json save, snapshot the xlsx and rebuild
            # the published HTML so the 'Wiki ↗' link is always fresh. Build
            # errors are surfaced via X-Build-Status header but don't fail the
            # save — authoring isn't blocked by a render bug.
            build_status = "skipped"
            build_msg = ""
            if target.name == "semantic.json":
                build_status, build_msg = self._rebuild_semantic()

            self.send_response(204)
            self.send_header("ETag", new_etag)
            self.send_header("X-Build-Status", build_status)
            if build_msg:
                # Headers must be single-line, ascii-ish; truncate just in case.
                safe = build_msg.replace("\n", " ").replace("\r", " ")[:500]
                self.send_header("X-Build-Message", safe)
            self.end_headers()
        except Exception as exc:  # pragma: no cover
            self.send_error(500, str(exc))

    def _rebuild_semantic(self) -> tuple[str, str]:
        """Invoke build_semantic.main([]) inline. Returns (status, message).
        Status is 'ok', 'error', or 'skipped'.

        We `importlib.reload()` build_semantic each time so edits to the
        generator's CSS / JS templates take effect without a server restart.
        Without this, the long-lived daemon serves whatever build_semantic
        code was loaded at first call — silently producing stale HTML.
        """
        try:
            import importlib
            import build_semantic  # type: ignore
            build_semantic = importlib.reload(build_semantic)
            # Capture stderr so build errors don't pollute server logs.
            from io import StringIO
            buf_out, buf_err = StringIO(), StringIO()
            old_out, old_err = sys.stdout, sys.stderr
            sys.stdout, sys.stderr = buf_out, buf_err
            try:
                code = build_semantic.main([])
            finally:
                sys.stdout, sys.stderr = old_out, old_err
            if code == 0:
                return "ok", "rebuilt 08_Semantic_Model.html"
            return "error", (buf_err.getvalue() or buf_out.getvalue()).strip()[:400]
        except Exception as exc:
            return "error", f"{type(exc).__name__}: {exc}"

    def do_POST(self):
        rel = self._path_no_query()
        # Auth endpoints (public — they ARE the login flow)
        if rel == "api/auth/login":
            return self._handle_auth_login()
        if rel == "api/auth/whoami":
            return self._handle_auth_whoami()
        if rel == "api/auth/change":
            return self._handle_auth_change()
        # All other POSTs (xlsx upload/export) require a valid bearer.
        if rel in _AUTH_POSTS and not _is_authed(self.headers):
            return self._send_401()
        if rel == "upload-inventory":
            return self._handle_upload_inventory()
        if rel == "upload-semantic":
            return self._handle_upload_semantic()
        if rel == "export-semantic-xlsx":
            return self._handle_export_semantic_xlsx()
        self.send_error(404, "no such endpoint")

    # ── Auth helpers ───────────────────────────────────────────────────
    def _send_json(self, code: int, payload: dict) -> None:
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(code)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def _send_401(self) -> None:
        self._send_json(401, {"error": "passcode required"})

    def _read_json_body(self) -> dict:
        length = int(self.headers.get("Content-Length", "0"))
        if length <= 0:
            return {}
        if length > 4096:
            raise ValueError("auth payload too large")
        raw = self.rfile.read(length)
        try:
            return json.loads(raw.decode("utf-8"))
        except Exception as exc:
            raise ValueError(f"invalid JSON body: {exc}")

    def _handle_auth_login(self) -> None:
        try:
            data = self._read_json_body()
        except ValueError as exc:
            return self._send_json(400, {"error": str(exc)})
        pin = str(data.get("passcode") or "")
        if not _verify_pin(pin):
            return self._send_json(401, {"error": "wrong passcode"})
        return self._send_json(200, {"token": _issue_token()})

    def _handle_auth_whoami(self) -> None:
        if _is_authed(self.headers):
            return self._send_json(200, {"ok": True})
        return self._send_401()

    def _handle_auth_change(self) -> None:
        if not _is_authed(self.headers):
            return self._send_401()
        try:
            data = self._read_json_body()
        except ValueError as exc:
            return self._send_json(400, {"error": str(exc)})
        cur = str(data.get("current") or "")
        new = str(data.get("new") or "")
        if not _verify_pin(cur):
            return self._send_json(401, {"error": "current passcode is wrong"})
        if not _PIN_RX.match(new):
            return self._send_json(400, {"error": "new passcode must be 6 digits"})
        if new == cur:
            return self._send_json(400, {"error": "new passcode must differ from current"})
        _set_pin(new)
        return self._send_json(200, {"token": _issue_token()})

    def _handle_upload_inventory(self):
        try:
            length = int(self.headers.get("Content-Length", "0"))
            if length <= 0:
                self.send_error(400, "empty request body")
                return
            if length > 20 * 1024 * 1024:
                self.send_error(413, "xlsx too large (max 20 MB)")
                return
            body = self.rfile.read(length)
            diff = _build_diff(body)
            payload = json.dumps(diff, ensure_ascii=False, default=str).encode("utf-8")
            self.send_response(200)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            self.send_header("Content-Length", str(len(payload)))
            self.end_headers()
            self.wfile.write(payload)
        except Exception as exc:
            msg = json.dumps({"error": str(exc)}).encode("utf-8")
            self.send_response(400)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            self.send_header("Content-Length", str(len(msg)))
            self.end_headers()
            self.wfile.write(msg)

    def _handle_upload_semantic(self):
        """One-shot xlsx → semantic.json. Overwrites the existing json
        (after rotating .bak). Used for first-run bootstrap and for explicit
        re-imports from the editor."""
        try:
            length = int(self.headers.get("Content-Length", "0"))
            if length <= 0:
                self.send_error(400, "empty request body")
                return
            if length > 20 * 1024 * 1024:
                self.send_error(413, "xlsx too large (max 20 MB)")
                return
            body = self.rfile.read(length)
            import build_semantic  # type: ignore
            # Write xlsx body to disk first (will replace whatever's there),
            # then call --import-xlsx in-process. Snapshot any existing xlsx
            # to backup/ before overwriting so we never silently lose edits.
            bkp = build_semantic._backup_xlsx()
            build_semantic.XLSX.parent.mkdir(parents=True, exist_ok=True)
            build_semantic.XLSX.write_bytes(body)
            data = build_semantic._read_xlsx()
            rpt = build_semantic.BuildReport()
            build_semantic._validate(data, rpt)
            if not rpt.ok:
                err = json.dumps({"error": "validation failed",
                                  "errors": rpt.errors,
                                  "warnings": rpt.warnings}).encode("utf-8")
                self.send_response(422)
                self.send_header("Content-Type", "application/json; charset=utf-8")
                self.send_header("Content-Length", str(len(err)))
                self.end_headers()
                self.wfile.write(err)
                return
            # Rotate existing semantic.json .bak before overwrite.
            jp = build_semantic.JSON_PATH
            if jp.exists():
                jp.with_suffix(jp.suffix + ".bak").write_bytes(jp.read_bytes())
            build_semantic._write_json(data)
            # Trigger rebuild so HTML reflects the new model.
            status, msg = self._rebuild_semantic()
            resp = {
                "ok": True,
                "xlsx_backup": bkp,
                "warnings": rpt.warnings,
                "build_status": status,
                "build_message": msg,
            }
            payload = json.dumps(resp, ensure_ascii=False).encode("utf-8")
            self.send_response(200)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            self.send_header("Content-Length", str(len(payload)))
            self.end_headers()
            self.wfile.write(payload)
        except Exception as exc:
            msg = json.dumps({"error": str(exc)}).encode("utf-8")
            self.send_response(400)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            self.send_header("Content-Length", str(len(msg)))
            self.end_headers()
            self.wfile.write(msg)

    def _handle_export_semantic_xlsx(self):
        """Snapshot the current xlsx to backup/, then write a fresh xlsx
        from semantic.json. No request body required."""
        try:
            import build_semantic  # type: ignore
            if not build_semantic.JSON_PATH.exists():
                self.send_error(409, "semantic.json does not exist yet")
                return
            data = build_semantic._read_json()
            bkp = build_semantic._backup_xlsx()
            build_semantic._write_xlsx(data)
            resp = {
                "ok": True,
                "xlsx_path": str(build_semantic.XLSX.relative_to(ROOT)),
                "xlsx_backup": bkp,
                "bytes": build_semantic.XLSX.stat().st_size,
            }
            payload = json.dumps(resp, ensure_ascii=False).encode("utf-8")
            self.send_response(200)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            self.send_header("Content-Length", str(len(payload)))
            self.end_headers()
            self.wfile.write(payload)
        except Exception as exc:
            msg = json.dumps({"error": str(exc)}).encode("utf-8")
            self.send_response(400)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            self.send_header("Content-Length", str(len(msg)))
            self.end_headers()
            self.wfile.write(msg)


def _kill_existing_on_port(port: int) -> None:
    """If another process is bound to `port`, kill it so we own the only instance.

    Loopback-only, stdlib-only. Skips silently if nothing is bound.
    """
    import socket, subprocess
    # Fast check: can we connect? If not, port is free — nothing to do.
    s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    s.settimeout(0.2)
    try:
        in_use = (s.connect_ex(("127.0.0.1", port)) == 0)
    finally:
        s.close()
    if not in_use:
        return

    me = os.getpid()
    killed = []
    try:
        if os.name == "nt":
            # netstat -ano | findstr LISTENING :PORT  → PID in last column
            out = subprocess.check_output(
                ["netstat", "-ano", "-p", "TCP"], text=True, stderr=subprocess.DEVNULL
            )
            pids = set()
            for line in out.splitlines():
                if f":{port} " in line and "LISTENING" in line:
                    parts = line.split()
                    if parts and parts[-1].isdigit():
                        pids.add(int(parts[-1]))
            for pid in pids:
                if pid == me:
                    continue
                subprocess.run(
                    ["taskkill", "/F", "/PID", str(pid)],
                    stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
                )
                killed.append(pid)
        else:
            # macOS / Linux: lsof -t -i tcp:PORT -sTCP:LISTEN
            out = subprocess.check_output(
                ["lsof", "-t", f"-iTCP:{port}", "-sTCP:LISTEN"],
                text=True, stderr=subprocess.DEVNULL,
            )
            for tok in out.split():
                if tok.isdigit() and int(tok) != me:
                    subprocess.run(["kill", "-9", tok],
                                   stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
                    killed.append(int(tok))
    except (subprocess.CalledProcessError, FileNotFoundError):
        pass

    if killed:
        print(f"  Killed previous serve_admin instance(s) on port {port}: PID {killed}")
        # Give Windows a brief moment to release the socket.
        import time; time.sleep(0.4)


def _parse_args(argv: list[str] | None = None) -> "argparse.Namespace":
    import argparse
    p = argparse.ArgumentParser(
        prog="serve_admin",
        description="Local handovers tooling server (admin.html + semantic_editor.html).",
    )
    grp = p.add_mutually_exclusive_group()
    grp.add_argument("--admin", dest="open_alias", action="store_const",
                     const="admin.html",
                     help="Open admin.html in the default browser (default).")
    grp.add_argument("--semantic", "--editor", dest="open_alias",
                     action="store_const", const="semantic_editor.html",
                     help="Open semantic_editor.html instead of admin.html.")
    grp.add_argument("--both", dest="open_alias", action="store_const",
                     const="BOTH",
                     help="Open both admin.html and semantic_editor.html.")
    grp.add_argument("--open", dest="open_path", metavar="PATH",
                     help="Open an arbitrary page (e.g. --open index.html).")
    grp.add_argument("--no-browser", dest="open_alias", action="store_const",
                     const="NONE",
                     help="Don't auto-open any page.")
    p.add_argument("--port", type=int, default=PORT,
                   help=f"TCP port to bind on 0.0.0.0 (default {PORT}).")
    p.add_argument("--bind", default="0.0.0.0",
                   help="Interface to bind to. Default 0.0.0.0 (all "
                        "interfaces, LAN-reachable). Use 127.0.0.1 to "
                        "restrict to loopback.")
    return p.parse_args(argv)


def _pages_to_open(ns) -> list[str]:
    """Map CLI args + env var to a list of relative paths to open."""
    # CLI > env > default. The legacy SERVE_ADMIN_NO_BROWSER kill-switch is
    # still honoured for back-compat with scripted callers.
    if os.environ.get("SERVE_ADMIN_NO_BROWSER"):
        return []
    if ns.open_alias == "NONE":
        return []
    if ns.open_path:
        return [ns.open_path.lstrip("/")]
    env_page = os.environ.get("SERVE_ADMIN_PAGE")
    if env_page:
        if env_page.upper() == "BOTH":
            return ["admin.html", "semantic_editor.html"]
        if env_page.upper() == "NONE":
            return []
        return [env_page.lstrip("/")]
    alias = ns.open_alias or "admin.html"
    if alias == "BOTH":
        return ["admin.html", "semantic_editor.html"]
    return [alias]


def main(argv: list[str] | None = None) -> None:
    ns = _parse_args(argv)
    port = ns.port
    bind = ns.bind
    # Make sure the passcode file exists so first-time users can log in.
    _load_passcode()
    _kill_existing_on_port(port)
    handler = partial(_Handler, directory=str(ROOT))
    # Bind per --bind (default 0.0.0.0 → LAN-reachable). The auth gate on
    # admin.html / semantic_editor.html / write endpoints is the protection.
    # ThreadingHTTPServer so a slow browser request can't block the PUT save.
    http.server.ThreadingHTTPServer.allow_reuse_address = True
    with http.server.ThreadingHTTPServer((bind, port), handler) as httpd:
        pages = _pages_to_open(ns)
        base = f"http://localhost:{port}"
        urls = [base] + [f"http://{ip}:{port}" for ip in _lan_ips()]
        print(f"Serving {ROOT}")
        for u in urls:
            print(f"  → {u}/")
        if pages:
            print(f"Opening {base}/{pages[0]}")
            for extra in pages[1:]:
                print(f"  + also opening {base}/{extra}")
        else:
            print("Auto-open disabled.")
        print("Press Ctrl-C to stop.")
        if pages:
            def _open_all():
                for i, p in enumerate(pages):
                    webbrowser.open(f"{base}/{p}", new=2)
                    if i + 1 < len(pages):
                        import time; time.sleep(0.4)
            threading.Timer(0.7, _open_all).start()
        try:
            httpd.serve_forever()
        except KeyboardInterrupt:
            print("\nStopped.")
            sys.exit(0)


if __name__ == "__main__":
    main()
